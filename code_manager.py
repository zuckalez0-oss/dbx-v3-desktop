# gerenciador dos códigos únicos das peças - desenvolvido para controle interno das peças.

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from app_paths import ensure_user_file

class CodeGenerator:
    """
    Gerencia a criação e persistência de códigos únicos para as peças.
    Interage com uma planilha Excel como banco de dados.
    """
    def __init__(self, db_path=None, error_notifier=None):
        self.db_path = db_path or ensure_user_file("codigo_database.xlsx")
        self.error_notifier = error_notifier
        self.code_column_name = 'Codigo Unico'
        self.timestamp_column_name = 'Data de Registro'
        self.project_column_name = 'Projeto'
        self.existing_codes = set()
        self.last_code_number = 0
        self.current_prefix = 'DES' # Prefixo padrão
        self._load_database()

    def _notify_error(self, title, message, level="warning"):
        if self.error_notifier:
            self.error_notifier(title, message, level)
            return
        print(f"{level.upper()}: {title}: {message}")

    def _get_custom_prefix(self):
        """Lê o prefixo personalizado da célula D2 (Linha 2, Coluna 4)."""
        try:
            wb = load_workbook(self.db_path, read_only=True, data_only=True)
            sheet = wb.active
            val = sheet.cell(row=2, column=4).value # D2
            wb.close()
            if val and str(val).strip():
                return str(val).strip()
        except Exception:
            pass
        return 'DES'

    def _load_database(self):
        self.current_prefix = self._get_custom_prefix()
        try:
            df = pd.read_excel(self.db_path)
            if self.code_column_name in df.columns:
                df[self.code_column_name] = df[self.code_column_name].astype(str)
                self.existing_codes = set(df[self.code_column_name].tolist())
                max_num = 0
                prefix_len = len(self.current_prefix)
                for code in self.existing_codes:
                    if code.upper().startswith(self.current_prefix.upper()):
                        try:
                            num = int(code[prefix_len:])
                            if num > max_num: max_num = num
                        except (ValueError, TypeError): continue
                self.last_code_number = max_num
        except FileNotFoundError:
            print(f"Arquivo '{self.db_path}' não encontrado. Um novo será criado.")
        except Exception as e:
            self._notify_error("Erro de Leitura", f"Não foi possível ler o banco de dados de códigos: {e}")

    def _append_to_database(self, new_code, project_number):
        timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        new_row = [new_code, timestamp, project_number]
        try:
            wb = load_workbook(self.db_path)
            sheet = wb.active
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
            sheet.append([self.code_column_name, self.timestamp_column_name, self.project_column_name])
        
        sheet.append(new_row)
        
        try:
            wb.save(self.db_path)
        except PermissionError:
             self._notify_error("Erro de Permissão", f"Não foi possível salvar o novo código em '{self.db_path}'.\n\nPor favor, feche a planilha e tente novamente.", level="critical")
             return False
        except Exception as e:
            self._notify_error("Erro ao Salvar", f"Ocorreu um erro ao salvar o banco de códigos: {e}", level="critical")
            return False
        return True

    def generate_new_code(self, project_number, prefix=None):
        self._load_database()
        effective_prefix = self.current_prefix
        current_number = self.last_code_number + 1
        new_code = f"{effective_prefix}{current_number}"
        while new_code in self.existing_codes:
            current_number += 1
            new_code = f"{effective_prefix}{current_number}"
        
        if self._append_to_database(new_code, project_number):
            self.existing_codes.add(new_code)
            self.last_code_number = current_number
            return new_code
        else:
            return None
