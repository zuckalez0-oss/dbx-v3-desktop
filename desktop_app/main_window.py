# versão 2.0.4  19/11/2025

from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook, Workbook
import sys
import os
import json
import shutil
import pandas as pd
import time
import unicodedata
from pathlib import Path
from app_paths import find_resource_path
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QLabel, QTextEdit, 
                             QFileDialog, QProgressBar, QMessageBox, QGroupBox,
                             QFormLayout, QLineEdit, QComboBox, QTableWidget, 
                             QTableWidgetItem, QDialog, QInputDialog, QHeaderView,
                             QSplitter, QDialogButtonBox, QSizePolicy, QScrollArea,
                             QFrame, QGridLayout, QTextBrowser) #importacao do pyqt5
from PyQt5.QtCore import Qt, QUrl
from PyQt5.QtGui import QIcon, QDesktopServices, QPixmap
try:
    import win32com.client
    import pythoncom
    PYWIN32_DISPONIVEL = True
except ImportError:
    PYWIN32_DISPONIVEL = False
    print("AVISO: 'pywin32' não encontrado. Geração de PDF do orçamento será pulada.")
    print("Para instalar, rode: pip install pywin32")
#classes encapsuladas em outros arquivos
from code_manager import CodeGenerator
from history_manager import HistoryManager
from .history_dialog import HistoryDialog
from .processing import ProcessThread
from .nesting_dialog import NestingDialog
from dxf_engine import get_dxf_bounding_box #importar funcao de dxf_engine
from calculo_cortes import orquestrar_planos_de_corte

LYPSYOS_WEBSITE_URL = os.environ.get("LYPSYOS_SITE_URL", "https://lypsyos.com")


def desktop_asset_path(*parts):
    base_dir = Path(__file__).resolve().parent
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        bundle_desktop_dir = Path(meipass) / "desktop_app"
        base_dir = bundle_desktop_dir if bundle_desktop_dir.exists() else Path(meipass)
    return str(base_dir.joinpath(*parts))


def first_existing_desktop_asset(*relative_paths):
    for relative_path in relative_paths:
        candidate = desktop_asset_path(*relative_path.split("/"))
        if os.path.exists(candidate):
            return candidate
    return ""


STYLE = """

/* Fundo principal e cor de texto padrão */
QWidget {
    background-color: #F7F7F7; /* Branco suave para o fundo principal */
    color: #173b5a;          /* Azul institucional Lypsyos */
    font-family: 'Segoe UI', Arial, sans-serif;
    font-size: 8.75pt; 
    border: none;
}

QLabel {
    min-height: 14px;
    line-height: 18px; 
    color: #173b5a;
    background: transparent;
    padding: 0 2px;
    font-weight: 500;
}

/* Divisores (Splitter) */
QSplitter::handle { background-color: #D5D8DC; }
QSplitter::handle:hover { background-color: #3498db; }
QSplitter::handle:pressed { background-color: #2980b9; }

QScrollArea {
    background: transparent;
    border: none;
}

QFrame#utilityBar {
    background-color: #FFFFFF;
    border: 1px solid #D7E3EA;
    border-radius: 9px;
}
QPushButton#utilityButton {
    background-color: transparent;
    color: #173b5a;
    border: 1px solid transparent;
    border-radius: 7px;
    padding: 3px 10px;
    min-height: 22px;
    max-height: 24px;
    font-size: 8.4pt;
}
QPushButton#utilityButton:hover {
    background-color: #EEF8F9;
    border: 1px solid #CDE8EA;
}
QPushButton#utilityButton:pressed {
    background-color: #DDEFF1;
}

/* Contêineres como GroupBox e Tabelas */
QGroupBox, QTableWidget, QListView {
    background-color: #FFFFFF; 
    border: 1px solid #D7E3EA; /* Borda cinza clara */
    border-radius: 8px;
}
QGroupBox {
    margin-top: 1.2em; 
    font-weight: bold;
    border: 1px solid #D7E3EA;
    border-radius: 8px;
    background-color: #FFFFFF;
    font-size: 9pt;
}

/* Títulos dos GroupBox */
QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top center;
    padding: 2px 12px;
    background-color: transparent;
    color: #1caeb8; /* Ciano institucional Lypsyos */
    border-radius: 5px;
    font-weight: bold;
}

/* Campos de Input e ComboBox */
QLineEdit, QTextEdit, QComboBox, QDoubleSpinBox, QSpinBox {
    background-color: #FFFFFF;
    border: 1px solid #D7E3EA;   /* Borda cinza clara */
    border-radius: 5px;
    padding: 1px 7px;
    min-height: 14px;       /* Altura mínima garantida no eixo Y */ 
    line-height: 18px;      /* Altura da linha para melhor centralização vertical */
    color: #173b5a;
    font-size: 8.75pt;      /* Aumento leve para melhor legibilidade */
}
QLineEdit:focus, QTextEdit:focus, QComboBox:focus, QDoubleSpinBox:focus, QSpinBox:focus,
QLineEdit:hover, QTextEdit:hover, QComboBox:hover, QDoubleSpinBox:hover, QSpinBox:hover {
    border: 1px solid #1ccfd4; /* Foco e hover com borda ciano */
}
QLineEdit::placeholder {
    color: #aab7c4;
}

/* Detalhes do ComboBox */
QComboBox::drop-down { border: none; }
QComboBox::down-arrow {
    /* Ícone SVG embutido para a seta, sem depender de arquivos externos */
    image: url('data:image/svg+xml;utf8,<svg xmlns="http://www.w3.org/2000/svg" width="12" height="12" viewBox="0 0 24 24"><path fill="%232c3e50" d="M7 10l5 5 5-5z"/></svg>');
    width: 12px; height: 12px; margin-right: 8px;
}
QComboBox QAbstractItemView {
    background-color: #FFFFFF;
    border: 1px solid #D7E3EA;
    selection-background-color: #1ccfd4;
    selection-color: #FFFFFF;
    outline: 0px;
}

/* Botões Padrão */
QPushButton {
    background-color: #EAF2F5;   /* neutro azulado */
    color: #173b5a;
    font-weight: bold;
    padding: 2px 7px; 
    min-height: 20px;
    max-height: 24px;
    border-radius: 6px;
    font-size: 8.25pt;
}
QPushButton:hover { background-color: #DCE9EF; }
QPushButton:pressed { background-color: #CBDDE6; }
QPushButton:disabled { background-color: #F3F6F8; color: #A8B7C2; }

/* Botão Primário (Ação principal) */
QPushButton#primaryButton { background-color: #2bd1d7; color: #12324b; }
QPushButton#primaryButton:hover { background-color: #20c0c6; }

/* Botões de estado */
QPushButton#successButton { background-color: #1f7d8c; color: #FFFFFF; }
QPushButton#successButton:hover { background-color: #17656f; }
QPushButton#warningButton { background-color: #173b5a; color: #FFFFFF; }
QPushButton#warningButton:hover { background-color: #102a40; }

/* Tabela */
QTableWidget {
    gridline-color: #EAECEE;
    border-radius: 4px;
}
QHeaderView::section {
    background-color: #F6FAFB;
    color: #173b5a;
    padding: 5px;
    border: none;
    border-bottom: 1px solid #D7E3EA;
    font-weight: bold;
}
QTableWidget::item {
    color: #173b5a;
    font-size: 9pt;
    padding: 5px;
    border-bottom: 1px solid #EAF2F5;
}
/* Linhas alternadas (zebradas) */
QTableWidget::item:alternate { background-color: #FDFEFE; }

/* Seleção da tabela */
QTableWidget::item:selected {
    background-color: #BDEEF0; /* Ciano claro */
    color: #173b5a;
}

/* Barra de Log */
QTextEdit#logExecution {
    font-family: 'Courier New', Courier, monospace;
    background-color: #EEF5F7;
    color: #173b5a;
    border-radius: 4px;
}

/* Barras de Rolagem */
QScrollBar:vertical { border: none; background: #F7F7F7; width: 10px; margin: 0; }
QScrollBar::handle:vertical { background: #BDC3C7; min-height: 20px; border-radius: 5px; }
QScrollBar::handle:vertical:hover { background: #95a5a6; }
QScrollBar:horizontal { border: none; background: #F7F7F7; height: 10px; margin: 0; }
QScrollBar::handle:horizontal { background: #BDC3C7; min-width: 20px; border-radius: 5px; }
QScrollBar::handle:horizontal:hover { background: #95a5a6; }
QScrollBar::add-line, QScrollBar::sub-line { border: none; background: none; }
"""

DARK = """
/* Fundo principal e cor de texto padrão */
QWidget {
    background-color: #173b5a; /* Azul institucional Lypsyos */
    color: #ECF0F1;          /* Cinza claro para o texto */
    font-family: 'Segoe UI', Arial, sans-serif;
    font-size: 8.75pt; 
    border: none;
}

QLabel {
    min-height: 14px;
    color: #ECF0F1;
    background: transparent;
    padding: 0 2px;
}

/* Divisores (Splitter) */
QSplitter::handle { background-color: #34495e; }
QSplitter::handle:hover { background-color: #3498db; }
QSplitter::handle:pressed { background-color: #2980b9; }

QScrollArea {
    background: transparent;
    border: none;
}

QFrame#utilityBar {
    background-color: #1d4662;
    border: 1px solid #2d5f78;
    border-radius: 9px;
}
QPushButton#utilityButton {
    background-color: transparent;
    color: #ECF0F1;
    border: 1px solid transparent;
    border-radius: 7px;
    padding: 3px 10px;
    min-height: 22px;
    max-height: 24px;
    font-size: 8.4pt;
}
QPushButton#utilityButton:hover {
    background-color: #244f6d;
    border: 1px solid #2d6f80;
}
QPushButton#utilityButton:pressed {
    background-color: #16384f;
}

/* Contêineres como GroupBox e Tabelas */
QGroupBox, QTableWidget, QListView {
    background-color: #34495e; 
    border: 1px solid #4a617a; /* Borda cinza-azulada */
    border-radius: 8px;
}
QGroupBox { margin-top: 1em; font-weight: bold; font-size: 9pt; }

/* Títulos dos GroupBox */
QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top center;
    padding: 2px 8px;
    background-color: transparent;
    color: #2bd1d7; /* Ciano institucional */
    border-radius: 4px;
    font-weight: bold;
}

QLineEdit, QTextEdit, QComboBox, QDoubleSpinBox, QSpinBox {
    background-color: #FFFFFF;
    border: 1px solid #D5D8DC;
    background-color: #2c3e50;
    border: 1px solid #4a617a;
    border-radius: 5px;
    
    /* SOLUÇÃO DO PROBLEMA DE ALTURA */
    min-height: 18px;       /* Altura mínima garantida no eixo Y */
    padding-left: 7px;      /* Respiro horizontal */
    padding-right: 7px;
    padding-top: 1px;       /* Centralização vertical visual */
    padding-bottom: 1px;
    
    font-size: 8.75pt;      /* Aumento leve para melhor legibilidade */
    color: #2c3e50;
    color: #ECF0F1;
}

/* Ajuste específico para o QComboBox não cortar o texto */
QComboBox {
    padding-left: 8px;
}

QLineEdit:focus, QComboBox:focus {
    border: 2px solid #2bd1d7; /* Borda mais espessa no foco para feedback visual */
    background-color: #FFFFFF;
    background-color: #2c3e50;
}

/* Detalhes do ComboBox */
QComboBox::drop-down { border: none; }
QComboBox::down-arrow {
    image: url('data:image/svg+xml;utf8,<svg xmlns="http://www.w3.org/2000/svg" width="12" height="12" viewBox="0 0 24 24"><path fill="%23ECF0F1" d="M7 10l5 5 5-5z"/></svg>');
    width: 12px; height: 12px; margin-right: 8px;
}
QComboBox QAbstractItemView {
    background-color: #34495e;
    border: 1px solid #4a617a;
    selection-background-color: #2bd1d7;
    selection-color: #FFFFFF;
    outline: 0px;
}

/* Botões Padrão */
QPushButton { background-color: #284d69; color: #ECF0F1; font-weight: bold; padding: 2px 7px; min-height: 20px; max-height: 24px; border-radius: 6px; font-size: 8.25pt; }
QPushButton:hover { background-color: #335d7d; }
QPushButton:pressed { background-color: #16324a; }
QPushButton:disabled { background-color: #3a4a5a; color: #7f8c8d; }

/* Botões especiais mantêm suas cores */
QPushButton#primaryButton { background-color: #2bd1d7; color: #12324b; }
QPushButton#primaryButton:hover { background-color: #20c0c6; }
QPushButton#successButton { background-color: #1f7d8c; color: #FFFFFF; }
QPushButton#successButton:hover { background-color: #17656f; }
QPushButton#warningButton { background-color: #2bd1d7; color: #12324b; }
QPushButton#warningButton:hover { background-color: #20c0c6; }

/* Tabela */
QTableWidget { gridline-color: #34495e; border-radius: 4px; }
QHeaderView::section { background-color: #2c3e50; color: #bdc3c7; padding: 6px; border: none; border-bottom: 1px solid #4a617a; font-weight: bold; }
QTableWidget::item { color: #bdc3c7; font-size: 9pt; padding: 6px; border-bottom: 1px solid #34495e; }
QTableWidget::item:alternate { background-color: #3a4a5a; }
QTableWidget::item:selected { background-color: #2980b9; color: #FFFFFF; }

/* Barra de Log */
QTextEdit#logExecution { font-family: 'Courier New', Courier, monospace; background-color: #222; color: #eee; border-radius: 4px; }

/* Barras de Rolagem */
QScrollBar:vertical { border: none; background: #2c3e50; width: 10px; margin: 0; }
QScrollBar::handle:vertical { background: #5c7590; min-height: 20px; border-radius: 5px; }
QScrollBar::handle:vertical:hover { background: #7f8c8d; }
QScrollBar:horizontal { border: none; background: #2c3e50; height: 10px; margin: 0; }
QScrollBar::handle:horizontal { background: #BDC3C7; min-width: 20px; border-radius: 5px; }
QScrollBar::handle:horizontal:hover { background: #95a5a6; }
QScrollBar::add-line, QScrollBar::sub-line { border: none; background: none; }
"""

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Gerador de Desenhos Técnicos e DXF - DBX V3")
        self.setGeometry(100, 100, 1280, 850) 
        self.setMinimumSize(1024, 720)
        self._apply_window_icon()

        self.code_generator = CodeGenerator(error_notifier=self._handle_code_generator_error)
        self.history_manager = HistoryManager()
        self.is_dark_theme = False 
        
        self.colunas_df = ['nome_arquivo', 'forma', 'espessura', 'qtd', 'largura', 'altura', 'diametro', 'rt_base', 'rt_height', 'trapezoid_large_base', 'trapezoid_small_base', 'trapezoid_height', 'furos']
        self.colunas_df = ['nome_arquivo', 'forma', 'espessura', 'qtd', 'largura', 'altura', 'diametro', 'rt_base', 'rt_height', 'trapezoid_large_base', 'trapezoid_small_base', 'trapezoid_height', 'furos', 'dxf_path']
        self.manual_df = pd.DataFrame(columns=self.colunas_df)
        self.excel_df = pd.DataFrame(columns=self.colunas_df)
        self.furos_atuais = []
        self.project_directory = None
        self.auto_upload_image_path = None
        self.auto_upload_json_path = None

        self.initUI() 
        self.connect_signals() 
        
        self.set_initial_button_state()
        self.update_dimension_fields(self.forma_combo.currentText())

    def _apply_window_icon(self):
        window_icon_path = first_existing_desktop_asset(
            "dbx-ly.ico",
            "dbx-ly.png",
            "dbx-ly.svg",
        )
        if window_icon_path:
            self.setWindowIcon(QIcon(window_icon_path))

    def _handle_code_generator_error(self, title, message, level="warning"):
        handlers = {
            "warning": QMessageBox.warning,
            "critical": QMessageBox.critical,
            "information": QMessageBox.information,
        }
        handlers.get(level, QMessageBox.warning)(self, title, message)

    def _create_form_label(self, text, min_width=126):
        label = QLabel(text)
        label.setMinimumWidth(min_width)
        label.setMaximumWidth(min_width + 8)
        label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        label.setWordWrap(False)
        return label

    def _normalize_text_key(self, value):
        normalized = unicodedata.normalize("NFKD", str(value or ""))
        normalized = normalized.encode("ascii", "ignore").decode("ascii")
        return normalized.strip().lower().replace("-", "_").replace(" ", "_")

    def _coerce_float(self, value, default=0.0):
        if value in (None, "", "-"):
            return default
        if isinstance(value, str):
            value = value.strip().replace(",", ".")
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    def _normalize_shape_value(self, value):
        shape_key = self._normalize_text_key(value)
        shape_map = {
            "rectangle": "rectangle",
            "rect": "rectangle",
            "retangulo": "rectangle",
            "retanguloo": "rectangle",
            "quadrado": "rectangle",
            "circle": "circle",
            "circulo": "circle",
            "right_triangle": "right_triangle",
            "righttriangle": "right_triangle",
            "triangulo_retangulo": "right_triangle",
            "triangulo": "right_triangle",
            "triangle": "right_triangle",
            "trapezoid": "trapezoid",
            "trapezio": "trapezoid",
            "dxf": "dxf_shape",
            "dxf_shape": "dxf_shape",
            "shape_dxf": "dxf_shape",
        }
        return shape_map.get(shape_key, "rectangle")

    def _normalize_holes_payload(self, raw_holes):
        if not isinstance(raw_holes, list):
            return []
        normalized_holes = []
        for hole in raw_holes:
            if not isinstance(hole, dict):
                continue
            diameter = self._coerce_float(
                hole.get("diam", hole.get("diametro", hole.get("diameter", 0.0)))
            )
            pos_x = self._coerce_float(hole.get("x", hole.get("pos_x", hole.get("x_pos", 0.0))))
            pos_y = self._coerce_float(hole.get("y", hole.get("pos_y", hole.get("y_pos", 0.0))))
            if diameter > 0:
                normalized_holes.append({"diam": diameter, "x": pos_x, "y": pos_y})
        return normalized_holes

    def _normalize_piece_payload(self, raw_piece, index):
        if not isinstance(raw_piece, dict):
            raise ValueError("registro de peça inválido")

        default_name = f"AI_PECA_{index:03d}"
        normalized_piece = {col: 0.0 for col in self.colunas_df}
        normalized_piece.update({"nome_arquivo": default_name, "forma": "rectangle", "furos": [], "dxf_path": ""})

        normalized_piece["nome_arquivo"] = str(
            raw_piece.get(
                "nome_arquivo",
                raw_piece.get("nome", raw_piece.get("id", raw_piece.get("codigo", default_name))),
            )
        ).strip() or default_name
        normalized_piece["forma"] = self._normalize_shape_value(
            raw_piece.get("forma", raw_piece.get("shape", raw_piece.get("tipo", "rectangle")))
        )
        normalized_piece["espessura"] = self._coerce_float(
            raw_piece.get("espessura", raw_piece.get("thickness", raw_piece.get("esp", 0.0)))
        )
        normalized_piece["qtd"] = self._coerce_float(
            raw_piece.get("qtd", raw_piece.get("quantidade", raw_piece.get("quantity", 1)))
        )
        normalized_piece["largura"] = self._coerce_float(
            raw_piece.get("largura", raw_piece.get("width", raw_piece.get("base", 0.0)))
        )
        normalized_piece["altura"] = self._coerce_float(
            raw_piece.get("altura", raw_piece.get("height", raw_piece.get("h", 0.0)))
        )
        normalized_piece["diametro"] = self._coerce_float(
            raw_piece.get("diametro", raw_piece.get("diameter", raw_piece.get("diam", 0.0)))
        )
        normalized_piece["rt_base"] = self._coerce_float(
            raw_piece.get("rt_base", raw_piece.get("base_triangulo", raw_piece.get("triangle_base", 0.0)))
        )
        normalized_piece["rt_height"] = self._coerce_float(
            raw_piece.get("rt_height", raw_piece.get("altura_triangulo", raw_piece.get("triangle_height", 0.0)))
        )
        normalized_piece["trapezoid_large_base"] = self._coerce_float(
            raw_piece.get("trapezoid_large_base", raw_piece.get("base_maior", raw_piece.get("large_base", 0.0)))
        )
        normalized_piece["trapezoid_small_base"] = self._coerce_float(
            raw_piece.get("trapezoid_small_base", raw_piece.get("base_menor", raw_piece.get("small_base", 0.0)))
        )
        normalized_piece["trapezoid_height"] = self._coerce_float(
            raw_piece.get("trapezoid_height", raw_piece.get("altura_trapezio", raw_piece.get("trap_height", 0.0)))
        )
        normalized_piece["furos"] = self._normalize_holes_payload(
            raw_piece.get("furos", raw_piece.get("holes", raw_piece.get("furacoes", [])))
        )
        normalized_piece["dxf_path"] = str(raw_piece.get("dxf_path", raw_piece.get("dxf", "")) or "").strip()
        return normalized_piece

    def _extract_pieces_from_json_payload(self, payload):
        if isinstance(payload, list):
            return payload
        if isinstance(payload, dict):
            if isinstance(payload.get("pieces"), list):
                return payload["pieces"]
            if isinstance(payload.get("pecas"), list):
                return payload["pecas"]
            if isinstance(payload.get("items"), list):
                return payload["items"]
        raise ValueError("o JSON precisa conter uma lista de peças em 'pieces', 'pecas' ou na raiz")

    def open_lypsyos_site(self):
        self.statusBar().showMessage("Abrindo site da Lypsyos...", 2000)
        if not QDesktopServices.openUrl(QUrl(LYPSYOS_WEBSITE_URL)):
            QMessageBox.warning(
                self,
                "Link indisponível",
                f"Não foi possível abrir o site configurado da Lypsyos:\n{LYPSYOS_WEBSITE_URL}",
            )

    def _build_about_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Sobre Nós")
        dialog.setMinimumSize(540, 360)
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(20, 18, 20, 18)
        layout.setSpacing(12)

        logo_label = QLabel()
        logo_label.setAlignment(Qt.AlignCenter)
        logo_path = first_existing_desktop_asset("dbx-ly.png", "lyps-v22-tm2-svg.png")
        if logo_path:
            pixmap = QPixmap(logo_path)
            if not pixmap.isNull():
                logo_label.setPixmap(pixmap.scaledToWidth(240, Qt.SmoothTransformation))
        layout.addWidget(logo_label)

        title_label = QLabel("Lypsyos")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 14pt; font-weight: 700; color: #1caeb8;")
        layout.addWidget(title_label)

        description = QTextBrowser()
        description.setOpenExternalLinks(False)
        description.setReadOnly(True)
        description.setMinimumHeight(150)
        description.setHtml(
            """
            <div style="font-size: 10pt; line-height: 1.55;">
                <p><b>DBX-V3</b> é a evolução da solução desktop voltada para preparação técnica, geração de DXF/PDF, organização de peças e apoio à produção.</p>
                <p>A plataforma foi desenvolvida pela <b>Lypsyos</b>, empresa focada em software e automação para a indústria, com ênfase em padronização, velocidade operacional e redução de retrabalho.</p>
                <p>Para conhecer mais sobre a empresa, serviços e soluções, acesse o site oficial:</p>
                <p><a href="https://lypsyos.com">https://lypsyos.com</a></p>
            </div>
            """
        )
        description.anchorClicked.connect(lambda _: self.open_lypsyos_site())
        layout.addWidget(description)

        actions_layout = QHBoxLayout()
        actions_layout.addStretch(1)
        open_site_btn = QPushButton("Abrir Site")
        open_site_btn.setObjectName("primaryButton")
        close_btn = QPushButton("Fechar")
        open_site_btn.clicked.connect(self.open_lypsyos_site)
        close_btn.clicked.connect(dialog.accept)
        actions_layout.addWidget(open_site_btn)
        actions_layout.addWidget(close_btn)
        layout.addLayout(actions_layout)
        return dialog

    def show_about_dialog(self):
        self._build_about_dialog().exec_()

    def show_help_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Help")
        dialog.setMinimumSize(700, 520)
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(18, 18, 18, 18)
        layout.setSpacing(12)

        title = QLabel("Passo a passo para usar a aplicação")
        title.setStyleSheet("font-size: 12pt; font-weight: 700; color: #1caeb8;")
        layout.addWidget(title)

        help_text = QTextBrowser()
        help_text.setReadOnly(True)
        help_text.setOpenExternalLinks(False)
        help_text.setHtml(
            """
            <div style="font-size: 10pt; line-height: 1.55;">
                <ol>
                    <li><b>Inicie um novo projeto</b><br/>Clique em <b>Iniciar Novo Projeto...</b> e defina o número/nome do projeto que será usado para salvar os arquivos.</li>
                    <li><b>Ajuste os parâmetros de custo</b><br/>Confira os campos de <b>Imposto</b> e <b>Frete</b> antes de exportar o resumo para Excel.</li>
                    <li><b>Baixe os arquivos base quando precisar</b><br/>Use o botão <b>Arquivos Base</b> no topo da janela para salvar a planilha modelo de produção, o arquivo <b>codigo_database.xlsx</b> e o guia rápido de uso.</li>
                    <li><b>Carregue dados externos, se necessário</b><br/>Use <b>Selecionar Planilha</b> para importar uma planilha de peças e <b>Importar DXF(s)</b> para trazer geometrias prontas.</li>
                    <li><b>Cadastre a peça manualmente</b><br/>Preencha <b>Nome/ID da Peça</b>, <b>Forma</b>, <b>Espessura</b>, <b>Quantidade</b> e as dimensões exibidas conforme o tipo escolhido.</li>
                    <li><b>Adicione furos quando aplicável</b><br/>Em <b>Furação Rápida</b>, replique furos em peças retangulares. Em <b>Furos Manuais</b>, informe diâmetro e coordenadas e clique em <b>Adicionar Furo Manual</b>.</li>
                    <li><b>Envie a peça para a lista</b><br/>Clique em <b>Adicionar Peça à Lista</b> para incluir o item na tabela de produção.</li>
                    <li><b>Revise a lista de produção</b><br/>Confira a grade à direita antes de gerar saídas. A lista reúne peças importadas e cadastradas manualmente.</li>
                    <li><b>Gere os arquivos finais</b><br/>Use os botões <b>Exportar Excel</b>, <b>Gerar PDF</b>, <b>Gerar DXF</b> ou <b>PDF + DXF</b> conforme a necessidade. Para cálculo de nesting, clique em <b>Aproveitamento</b>.</li>
                    <li><b>Finalize o projeto</b><br/>Quando o lote estiver validado, clique em <b>Projeto Concluído</b> para registrar o histórico e limpar a sessão atual.</li>
                    <li><b>Consulte o log</b><br/>O painel <b>Log de Execução</b> mostra erros, avisos e o andamento de exportações e processamentos.</li>
                </ol>
                <p><b>Dica:</b> use <b>Ver Histórico de Projetos</b> para reabrir trabalhos anteriores sem precisar reconstruir a sessão manualmente.</p>
            </div>
            """
        )
        layout.addWidget(help_text, 1)

        close_layout = QHBoxLayout()
        close_layout.addStretch(1)
        close_btn = QPushButton("Fechar")
        close_btn.clicked.connect(dialog.accept)
        close_layout.addWidget(close_btn)
        layout.addLayout(close_layout)
        dialog.exec_()

    def _default_downloads_dir(self):
        downloads_dir = Path.home() / "Downloads"
        if downloads_dir.exists():
            return downloads_dir
        return Path.home()

    def _support_file_source(self, filename):
        source_path = find_resource_path(filename)
        if source_path is None:
            raise FileNotFoundError(f"O arquivo de suporte '{filename}' nao foi encontrado no pacote da aplicacao.")
        return Path(source_path)

    def _support_instructions_text(self):
        active_db_dir = Path(self.code_generator.db_path).parent
        return "\n".join(
            [
                "ARQUIVOS BASE DBX-V3",
                "",
                "1. planilha-dbx.xlsx",
                "- Use esta planilha como layout modelo para importar pecas em lote.",
                "- Preencha uma linha por peca e salve o arquivo.",
                "- Na aplicacao, clique em 'Selecionar Planilha' para importar o lote.",
                "- Colunas comuns: nome_arquivo, forma, espessura, qtd, largura, altura, diametro e furos.",
                "- Para furos em lote, use grupos como: furo_1_diametro, furo_1_x, furo_1_y.",
                "",
                "2. codigo_database.xlsx",
                "- Esse arquivo controla a base de codigos usada no botao 'Gerar Codigo'.",
                "- Para usar prefixo personalizado, edite a celula D2 com o prefixo desejado.",
                "- Exemplo de prefixo: LAS, CNC, PRT ou qualquer sigla padrao da sua operacao.",
                "- Salve o arquivo e feche o Excel antes de instalar/substituir a base.",
                "- Depois, na aplicacao, abra 'Arquivos Base' e use 'Instalar Codigo Database'.",
                "",
                "3. Pasta ativa da aplicacao",
                f"- Local atual da base ativa: {active_db_dir}",
                "",
                "DICAS",
                "- O pacote completo exporta os dois arquivos e este guia em uma pasta unica.",
                "- Se o Excel estiver aberto durante a instalacao da base, o Windows pode bloquear a copia.",
                "- Apos instalar uma nova base de codigos, o prefixo passa a valer nos proximos codigos gerados.",
            ]
        )

    def _copy_support_file(self, filename, destination_path):
        source_path = self._support_file_source(filename)
        destination = Path(destination_path)
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_path, destination)
        return destination

    def export_support_file(self, filename, label):
        try:
            default_path = self._default_downloads_dir() / filename
            save_path, _ = QFileDialog.getSaveFileName(
                self,
                f"Salvar {label}",
                str(default_path),
                "Excel Files (*.xlsx)",
            )
            if not save_path:
                return

            target_path = Path(save_path)
            if target_path.suffix.lower() != ".xlsx":
                target_path = target_path.with_suffix(".xlsx")

            copied_path = self._copy_support_file(filename, target_path)
            self.log_text.append(f"Arquivo de suporte exportado: {copied_path}")
            self.statusBar().showMessage(f"{label} salvo com sucesso.", 3000)
            QMessageBox.information(
                self,
                "Arquivo salvo",
                f"{label} salvo com sucesso em:\n{copied_path}",
            )
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Falha ao salvar arquivo",
                f"Nao foi possivel salvar '{filename}':\n{exc}",
            )

    def export_support_package(self):
        try:
            base_dir = QFileDialog.getExistingDirectory(
                self,
                "Escolher pasta para salvar o pacote de suporte",
                str(self._default_downloads_dir()),
            )
            if not base_dir:
                return

            package_dir = Path(base_dir) / "DBX-V3_Arquivos_Base"
            package_dir.mkdir(parents=True, exist_ok=True)

            copied_files = [
                self._copy_support_file("planilha-dbx.xlsx", package_dir / "planilha-dbx.xlsx"),
                self._copy_support_file("codigo_database.xlsx", package_dir / "codigo_database.xlsx"),
            ]
            instructions_path = package_dir / "INSTRUCOES_DBX-V3.txt"
            instructions_path.write_text(self._support_instructions_text(), encoding="utf-8")

            self.log_text.append(f"Pacote de suporte exportado em: {package_dir}")
            self.statusBar().showMessage("Pacote de suporte salvo com sucesso.", 3000)
            QMessageBox.information(
                self,
                "Pacote salvo",
                "Os arquivos base foram exportados com sucesso.\n\n"
                f"Pasta: {package_dir}\n"
                f"Itens: {len(copied_files)} planilhas + guia de instrucoes.",
            )
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Falha ao exportar pacote",
                f"Nao foi possivel exportar os arquivos base:\n{exc}",
            )

    def install_code_database_from_file(self):
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Selecionar codigo_database.xlsx personalizado",
                str(self._default_downloads_dir()),
                "Excel Files (*.xlsx)",
            )
            if not file_path:
                return

            destination = Path(self.code_generator.db_path)
            source = Path(file_path)
            destination.parent.mkdir(parents=True, exist_ok=True)

            if source.resolve() != destination.resolve():
                shutil.copy2(source, destination)

            self.code_generator._load_database()
            self.log_text.append(
                f"Base de codigos instalada/atualizada em '{destination}' com prefixo atual '{self.code_generator.current_prefix}'."
            )
            self.statusBar().showMessage("Codigo database instalado com sucesso.", 3000)
            QMessageBox.information(
                self,
                "Base instalada",
                "O arquivo de codigos foi atualizado com sucesso.\n\n"
                f"Prefixo atual detectado: {self.code_generator.current_prefix}\n"
                f"Local ativo: {destination}",
            )
        except PermissionError:
            QMessageBox.critical(
                self,
                "Arquivo em uso",
                "Nao foi possivel instalar o codigo_database.xlsx porque ele esta aberto em outro programa.\n\n"
                "Feche o Excel e tente novamente.",
            )
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Falha na instalacao",
                f"Nao foi possivel instalar o codigo_database.xlsx:\n{exc}",
            )

    def open_support_data_folder(self):
        folder_path = str(Path(self.code_generator.db_path).parent)
        self.statusBar().showMessage("Abrindo pasta de dados da aplicacao...", 2500)
        if not QDesktopServices.openUrl(QUrl.fromLocalFile(folder_path)):
            QMessageBox.warning(
                self,
                "Pasta indisponivel",
                f"Nao foi possivel abrir a pasta de dados:\n{folder_path}",
            )

    def show_support_files_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Arquivos Base")
        dialog.setMinimumSize(760, 540)
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(18, 18, 18, 18)
        layout.setSpacing(12)

        title = QLabel("Baixar arquivos de suporte da aplicacao")
        title.setStyleSheet("font-size: 12pt; font-weight: 700; color: #1caeb8;")
        layout.addWidget(title)

        intro = QLabel(
            "Use este painel para baixar os modelos base da aplicacao, instalar uma base de codigos personalizada e consultar o modo correto de uso."
        )
        intro.setWordWrap(True)
        layout.addWidget(intro)

        active_path_label = QLabel(f"Pasta ativa da aplicacao: {Path(self.code_generator.db_path).parent}")
        active_path_label.setWordWrap(True)
        active_path_label.setStyleSheet("font-size: 8.4pt;")
        layout.addWidget(active_path_label)

        content = QTextBrowser()
        content.setReadOnly(True)
        content.setOpenExternalLinks(False)
        content.setHtml(
            """
            <div style="font-size: 10pt; line-height: 1.58;">
                <p><b>Arquivos disponiveis:</b></p>
                <ul>
                    <li><b>planilha-dbx.xlsx</b><br/>Modelo para cadastro de pecas em lote. Preencha a planilha, salve e depois use <b>Selecionar Planilha</b> na tela principal.</li>
                    <li><b>codigo_database.xlsx</b><br/>Base de codigos usada pelo botao <b>Gerar Codigo</b>. Para definir um prefixo personalizado, edite a celula <b>D2</b>, salve o arquivo e depois use <b>Instalar Codigo Database</b>.</li>
                </ul>
                <p><b>Fluxo recomendado:</b></p>
                <ol>
                    <li>Clique em <b>Baixar Pacote Completo</b> para receber os dois arquivos e o guia rapido.</li>
                    <li>Edite a <b>planilha-dbx.xlsx</b> para montar o lote de pecas.</li>
                    <li>Se quiser prefixo proprio, edite <b>D2</b> no <b>codigo_database.xlsx</b>.</li>
                    <li>Com o Excel fechado, use <b>Instalar Codigo Database</b> para ativar a base personalizada.</li>
                </ol>
                <p><b>Observacao:</b> manter o arquivo aberto no Excel pode bloquear a instalacao ou a substituicao da base.</p>
            </div>
            """
        )
        layout.addWidget(content, 1)

        actions_grid = QGridLayout()
        actions_grid.setHorizontalSpacing(8)
        actions_grid.setVerticalSpacing(8)

        download_package_btn = QPushButton("Baixar Pacote Completo")
        download_package_btn.setObjectName("primaryButton")
        download_sheet_btn = QPushButton("Baixar Planilha de Produção")
        download_db_btn = QPushButton("Baixar Código Database")
        install_db_btn = QPushButton("Instalar Código Database")
        open_data_btn = QPushButton("Abrir Pasta de Dados")
        close_btn = QPushButton("Fechar")

        download_package_btn.clicked.connect(self.export_support_package)
        download_sheet_btn.clicked.connect(lambda: self.export_support_file("planilha-dbx.xlsx", "Planilha de Producao"))
        download_db_btn.clicked.connect(lambda: self.export_support_file("codigo_database.xlsx", "Codigo Database"))
        install_db_btn.clicked.connect(self.install_code_database_from_file)
        open_data_btn.clicked.connect(self.open_support_data_folder)
        close_btn.clicked.connect(dialog.accept)

        actions_grid.addWidget(download_package_btn, 0, 0)
        actions_grid.addWidget(download_sheet_btn, 0, 1)
        actions_grid.addWidget(download_db_btn, 1, 0)
        actions_grid.addWidget(install_db_btn, 1, 1)
        actions_grid.addWidget(open_data_btn, 2, 0)
        actions_grid.addWidget(close_btn, 2, 1)
        layout.addLayout(actions_grid)

        dialog.exec_()

    def show_v3_features_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Novidades DBX-V3")
        dialog.setMinimumSize(760, 520)
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(18, 18, 18, 18)
        layout.setSpacing(12)

        title = QLabel("Escalonamento da plataforma para DBX-V3")
        title.setStyleSheet("font-size: 12pt; font-weight: 700; color: #1caeb8;")
        layout.addWidget(title)

        content = QTextBrowser()
        content.setReadOnly(True)
        content.setOpenExternalLinks(False)
        content.setHtml(
            """
            <div style="font-size: 10pt; line-height: 1.6;">
                <p><b>DBX-V3</b> amplia a base da aplicação desktop para suportar entrada de dados mais dinâmica, padronizada e escalável.</p>
                <p><b>Novas direções já preparadas nesta versão:</b></p>
                <ul>
                    <li><b>Sessão Upload Automático</b> para importar JSON de peças gerado por IA ou por integrações externas.</li>
                    <li><b>Fluxo de imagem para JSON</b> preparado para receber desenhos, croquis e imagens técnicas que serão processados por um modelo especialista.</li>
                    <li><b>Estrutura de normalização</b> para converter JSON em peças válidas com forma, espessura, quantidade, dimensões e furos.</li>
                    <li><b>UI preparada para expansão</b> sem quebrar o fluxo atual de cadastro manual, planilha e DXF.</li>
                </ul>
                <p><b>Estratégia do fluxo IA:</b></p>
                <ol>
                    <li>O usuário envia uma imagem, PDF ou desenho manual.</li>
                    <li>Uma IA especialista em leitura de peças devolve um <b>JSON estruturado</b>.</li>
                    <li>A aplicação normaliza esse JSON e injeta as peças diretamente na lista de produção.</li>
                </ol>
                <p>Nesta etapa, a <b>importação de JSON</b> já está integrada ao desktop. O processamento visual por IA fica preparado como a próxima camada de evolução.</p>
            </div>
            """
        )
        layout.addWidget(content, 1)

        actions = QHBoxLayout()
        actions.addStretch(1)
        schema_btn = QPushButton("Ver Modelo JSON")
        schema_btn.setObjectName("primaryButton")
        close_btn = QPushButton("Fechar")
        schema_btn.clicked.connect(self.show_upload_schema_dialog)
        close_btn.clicked.connect(dialog.accept)
        actions.addWidget(schema_btn)
        actions.addWidget(close_btn)
        layout.addLayout(actions)
        dialog.exec_()

    def show_upload_schema_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Modelo JSON - Upload Automático")
        dialog.setMinimumSize(760, 540)
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(18, 18, 18, 18)
        layout.setSpacing(12)

        title = QLabel("Estrutura recomendada para JSON de peças")
        title.setStyleSheet("font-size: 12pt; font-weight: 700; color: #1caeb8;")
        layout.addWidget(title)

        schema_view = QTextBrowser()
        schema_view.setReadOnly(True)
        schema_view.setOpenExternalLinks(False)
        schema_view.setPlainText(
            json.dumps(
                {
                    "source_image": "croqui_lote_01.png",
                    "pieces": [
                        {
                            "nome_arquivo": "CEVA021",
                            "forma": "rectangle",
                            "espessura": 2.0,
                            "qtd": 10,
                            "largura": 200.0,
                            "altura": 200.0,
                            "furos": [
                                {"diam": 15.0, "x": 35.0, "y": 35.0},
                                {"diam": 15.0, "x": 165.0, "y": 35.0},
                                {"diam": 15.0, "x": 165.0, "y": 165.0},
                                {"diam": 15.0, "x": 35.0, "y": 165.0},
                            ],
                        },
                        {
                            "nome_arquivo": "TRI_002",
                            "forma": "right_triangle",
                            "espessura": 3.0,
                            "qtd": 2,
                            "rt_base": 140.0,
                            "rt_height": 80.0,
                            "furos": [],
                        },
                    ],
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        layout.addWidget(schema_view, 1)

        note = QLabel("Campos equivalentes como 'shape', 'quantity', 'width', 'height', 'holes' e 'diameter' também são aceitos na importação.")
        note.setWordWrap(True)
        layout.addWidget(note)

        footer = QHBoxLayout()
        footer.addStretch(1)
        close_btn = QPushButton("Fechar")
        close_btn.clicked.connect(dialog.accept)
        footer.addWidget(close_btn)
        layout.addLayout(footer)
        dialog.exec_()

    def select_auto_upload_image(self):
        if not self.project_directory:
            QMessageBox.warning(self, "Nenhum Projeto Ativo", "Inicie um projeto antes de usar a sessão de upload automático.")
            return

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Selecionar imagem ou desenho base",
            "",
            "Arquivos Compatíveis (*.png *.jpg *.jpeg *.bmp *.webp *.pdf)",
        )
        if not file_path:
            return

        self.auto_upload_image_path = file_path
        self.upload_status_label.setText(
            f"Base selecionada: {os.path.basename(file_path)} | Próximo passo: gerar JSON especializado e importar."
        )
        self.log_text.append(f"Upload automático: base visual selecionada '{os.path.basename(file_path)}'.")

    def import_auto_json(self):
        if not self.project_directory:
            QMessageBox.warning(self, "Nenhum Projeto Ativo", "Inicie um projeto antes de importar JSON automático.")
            return

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Importar JSON de peças",
            "",
            "JSON Files (*.json)",
        )
        if not file_path:
            return

        try:
            with open(file_path, "r", encoding="utf-8") as json_file:
                payload = json.load(json_file)

            raw_pieces = self._extract_pieces_from_json_payload(payload)
            normalized_pieces = []
            skipped_records = []
            for index, raw_piece in enumerate(raw_pieces, start=1):
                try:
                    normalized_pieces.append(self._normalize_piece_payload(raw_piece, index))
                except Exception as exc:
                    skipped_records.append(f"Registro {index}: {exc}")

            if not normalized_pieces:
                raise ValueError("nenhuma peça válida foi encontrada no JSON importado")

            new_df = pd.DataFrame(normalized_pieces, columns=self.colunas_df)
            self.manual_df = pd.concat([self.manual_df, new_df], ignore_index=True)
            self.auto_upload_json_path = file_path
            self.upload_status_label.setText(
                f"JSON importado: {os.path.basename(file_path)} | {len(normalized_pieces)} peça(s) adicionada(s)."
            )
            self.log_text.append(
                f"Upload automático: JSON '{os.path.basename(file_path)}' importado com {len(normalized_pieces)} peça(s)."
            )
            if skipped_records:
                for skipped in skipped_records:
                    self.log_text.append(f"Upload automático - aviso: {skipped}")

            self.update_table_display()
            self.set_initial_button_state()
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Falha na importação automática",
                f"Não foi possível importar o JSON informado:\n{exc}",
            )

    def initUI(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(14, 10, 14, 14)
        main_layout.setSpacing(10)

        utility_bar = QFrame()
        utility_bar.setObjectName("utilityBar")
        utility_layout = QHBoxLayout(utility_bar)
        utility_layout.setContentsMargins(10, 6, 10, 6)
        utility_layout.setSpacing(6)
        utility_title = QLabel("DBX-V3")
        utility_title.setStyleSheet("font-size: 9pt; font-weight: 700; color: #1caeb8;")
        self.support_files_btn = QPushButton("Arquivos Base")
        self.v3_features_btn = QPushButton("Novidades V3")
        self.about_btn = QPushButton("Sobre Nós")
        self.help_btn = QPushButton("Help")
        for button, width in [
            (self.support_files_btn, 118),
            (self.v3_features_btn, 112),
            (self.about_btn, 92),
            (self.help_btn, 72),
        ]:
            button.setObjectName("utilityButton")
            button.setMinimumWidth(width)
            button.setMaximumWidth(width + 8)
            button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        utility_layout.addWidget(utility_title, 0, Qt.AlignVCenter)
        utility_layout.addStretch(1)
        utility_layout.addWidget(self.support_files_btn)
        utility_layout.addWidget(self.v3_features_btn)
        utility_layout.addWidget(self.about_btn)
        utility_layout.addWidget(self.help_btn)
        main_layout.addWidget(utility_bar, 0)

        top_h_layout = QHBoxLayout()
        top_h_layout.setContentsMargins(0, 0, 0, 0)
        top_h_layout.setSpacing(12)
        top_h_layout.setAlignment(Qt.AlignTop)
        
    
        left_panel_widget = QWidget()
        left_v_layout = QVBoxLayout(left_panel_widget)
        left_v_layout.setContentsMargins(0,0,0,0)
        left_v_layout.setSpacing(6)
        left_panel_widget.setMinimumWidth(640)
        left_panel_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)


        project_group = QGroupBox("1. Projeto")
        project_layout = QHBoxLayout()
        project_layout.setContentsMargins(8, 7, 8, 6)
        project_layout.setSpacing(4)
        self.start_project_btn = QPushButton("Iniciar Novo Projeto...")
        self.theme_toggle_btn = QPushButton("🌙 Tema Escuro")
        self.history_btn = QPushButton("Ver Histórico de Projetos")
        top_buttons = [
            (self.start_project_btn, 168),
            (self.theme_toggle_btn, 138),
            (self.history_btn, 178),
        ]
        for button, max_width in top_buttons:
            button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            button.setMinimumWidth(max_width - 18)
            button.setMaximumWidth(max_width)
            button.setMinimumHeight(22)
            button.setMaximumHeight(24)
        project_layout.addWidget(self.start_project_btn)
        project_layout.addWidget(self.theme_toggle_btn)
        project_layout.addWidget(self.history_btn)
        project_layout.addStretch(1)
        project_group.setLayout(project_layout)
        project_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Maximum)
        left_v_layout.addWidget(project_group)
        

        cost_group = QGroupBox("2. Parâmetros de Custo")
        cost_layout = QFormLayout()
        cost_layout.setContentsMargins(10, 10, 10, 8)
        cost_layout.setHorizontalSpacing(8)
        cost_layout.setVerticalSpacing(4)

        self.imposto_input = QLineEdit("0,12") 
        self.frete_input = QLineEdit("0,26")

        self.imposto_input.setFixedWidth(110)
        self.frete_input.setFixedWidth(110)
        cost_layout.addRow(self._create_form_label("Imposto (%):", 88), self.imposto_input)
        cost_layout.addRow(self._create_form_label("Frete (R$):", 88), self.frete_input)
        cost_group.setLayout(cost_layout)
        cost_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Maximum)
        left_v_layout.addWidget(cost_group)

        file_group = QGroupBox("3. Carregar Planilha (Opcional)")
        file_layout = QVBoxLayout()
        file_layout.setContentsMargins(10, 9, 10, 8)
        file_layout.setSpacing(5)
        self.file_label = QLabel("Nenhum projeto ativo.")
        file_button_layout = QHBoxLayout()
        file_button_layout.setSpacing(4)
        self.select_file_btn = QPushButton("Selecionar Planilha")
        self.import_dxf_btn = QPushButton("Importar DXF(s)")
        self.clear_excel_btn = QPushButton("Limpar Planilha")
        for button in [self.select_file_btn, self.import_dxf_btn, self.clear_excel_btn]:
            button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        file_button_layout.addWidget(self.select_file_btn)
        file_button_layout.addWidget(self.import_dxf_btn)
        file_button_layout.addWidget(self.clear_excel_btn)
        file_layout.addWidget(self.file_label)
        file_layout.addLayout(file_button_layout)
        file_group.setLayout(file_layout)
        file_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Maximum)
        left_v_layout.addWidget(file_group)

        upload_group = QGroupBox("Sessão Upload Automático")
        upload_layout = QVBoxLayout()
        upload_layout.setContentsMargins(10, 9, 10, 8)
        upload_layout.setSpacing(5)
        self.upload_hint_label = QLabel(
            "DBX-V3: selecione uma imagem/croqui, gere um JSON com IA especialista e importe as peças automaticamente."
        )
        self.upload_hint_label.setWordWrap(True)
        self.upload_hint_label.setStyleSheet("font-size: 8.4pt; color: #607080;")
        self.upload_status_label = QLabel("Nenhuma base visual ou JSON automático selecionado.")
        self.upload_status_label.setWordWrap(True)
        self.upload_status_label.setStyleSheet("font-size: 8.4pt; font-style: italic; color: #607080;")
        upload_buttons_layout = QHBoxLayout()
        upload_buttons_layout.setContentsMargins(0, 0, 0, 0)
        upload_buttons_layout.setSpacing(4)
        self.select_auto_image_btn = QPushButton("Selecionar Imagem")
        self.import_auto_json_btn = QPushButton("Importar JSON")
        self.upload_schema_btn = QPushButton("Modelo JSON")
        for button in [self.select_auto_image_btn, self.import_auto_json_btn, self.upload_schema_btn]:
            button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            button.setMinimumHeight(22)
            button.setMaximumHeight(24)
        upload_buttons_layout.addWidget(self.select_auto_image_btn)
        upload_buttons_layout.addWidget(self.import_auto_json_btn)
        upload_buttons_layout.addWidget(self.upload_schema_btn)
        upload_layout.addWidget(self.upload_hint_label)
        upload_layout.addWidget(self.upload_status_label)
        upload_layout.addLayout(upload_buttons_layout)
        upload_group.setLayout(upload_layout)
        upload_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Maximum)
        left_v_layout.addWidget(upload_group)

        manual_group = QGroupBox("4. Informações da Peça")
        manual_layout = QFormLayout()
        manual_layout.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter)
        manual_layout.setFormAlignment(Qt.AlignTop)
        manual_layout.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        manual_layout.setRowWrapPolicy(QFormLayout.DontWrapRows)
        manual_layout.setVerticalSpacing(4)
        manual_layout.setHorizontalSpacing(8)
        manual_layout.setContentsMargins(10, 10, 10, 8)
        self.projeto_input = QLineEdit()
        self.projeto_input.setReadOnly(True)
        self.projeto_input.setMinimumWidth(260)
        manual_layout.addRow(self._create_form_label("Nº do Projeto Ativo:", 122), self.projeto_input)
        self.nome_input = QLineEdit()
        self.generate_code_btn = QPushButton("Gerar Código")
        self.generate_code_btn.setMinimumWidth(84)
        self.generate_code_btn.setMaximumWidth(94)
        self.generate_code_btn.setMinimumHeight(22)
        self.generate_code_btn.setMaximumHeight(24)
        self.generate_code_btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        name_layout = QHBoxLayout()
        name_layout.setContentsMargins(0, 0, 0, 0)
        # O stretch=1 faz com que o campo de nome ocupe o espaço extra, impedindo o botão de esticar
        name_layout.addWidget(self.nome_input, 1)
        name_layout.addWidget(self.generate_code_btn)
        name_layout.setSpacing(3)
        self.nome_input.setMinimumWidth(280)
        manual_layout.addRow(self._create_form_label("Nome/ID da Peça:", 122), name_layout)
        self.forma_combo = QComboBox()
        self.forma_combo.addItems(['rectangle', 'circle', 'right_triangle', 'trapezoid', 'dxf_shape'])
        self.espessura_input, self.qtd_input = QLineEdit(), QLineEdit()
        manual_layout.addRow(self._create_form_label("Forma:", 122), self.forma_combo)
        manual_layout.addRow(self._create_form_label("Espessura (mm):", 122), self.espessura_input)
        manual_layout.addRow(self._create_form_label("Quantidade:", 122), self.qtd_input)
        self.largura_input, self.altura_input = QLineEdit(), QLineEdit()
        self.diametro_input, self.rt_base_input, self.rt_height_input = QLineEdit(), QLineEdit(), QLineEdit()
        self.trapezoid_large_base_input, self.trapezoid_small_base_input, self.trapezoid_height_input = QLineEdit(), QLineEdit(), QLineEdit()
        self.largura_row = [self._create_form_label("Largura:", 122), self.largura_input]; manual_layout.addRow(*self.largura_row)
        self.altura_row = [self._create_form_label("Altura:", 122), self.altura_input]; manual_layout.addRow(*self.altura_row)
        self.diametro_row = [self._create_form_label("Diâmetro:", 122), self.diametro_input]; manual_layout.addRow(*self.diametro_row)
        self.rt_base_row = [self._create_form_label("Base Triângulo:", 122), self.rt_base_input]; manual_layout.addRow(*self.rt_base_row)
        self.rt_height_row = [self._create_form_label("Altura Triângulo:", 122), self.rt_height_input]; manual_layout.addRow(*self.rt_height_row)
        self.trap_large_base_row = [self._create_form_label("Base Maior:", 122), self.trapezoid_large_base_input]; manual_layout.addRow(*self.trap_large_base_row)
        self.trap_small_base_row = [self._create_form_label("Base Menor:", 122), self.trapezoid_small_base_input]; manual_layout.addRow(*self.trap_small_base_row)
        self.trap_height_row = [self._create_form_label("Altura Trapézio:", 122), self.trapezoid_height_input]; manual_layout.addRow(*self.trap_height_row)
        self.add_piece_btn = QPushButton("Adicionar Peça à Lista")
        self.add_piece_btn.setObjectName("primaryButton")
        self.add_piece_btn.setMinimumHeight(24)
        self.add_piece_btn.setMaximumHeight(26)
        manual_layout.addRow(self.add_piece_btn)
        manual_group.setLayout(manual_layout)
        manual_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Maximum)
        self.manual_group = manual_group
        self.manual_scroll_area = QScrollArea()
        self.manual_scroll_area.setWidgetResizable(True)
        self.manual_scroll_area.setFrameShape(QFrame.NoFrame)
        self.manual_scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.manual_scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.manual_scroll_area.setMinimumHeight(286)
        self.manual_scroll_area.setWidget(self.manual_group)
        self.manual_scroll_area.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        left_v_layout.addWidget(self.manual_scroll_area, 1)
        
        furos_main_group = QGroupBox("5. Adicionar Furos")
        furos_main_layout = QVBoxLayout()
        furos_main_layout.setContentsMargins(10, 10, 10, 8)
        furos_main_layout.setSpacing(5)
        self.rep_group = QGroupBox("Furação Rápida")
        rep_layout = QFormLayout()
        rep_layout.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        rep_layout.setRowWrapPolicy(QFormLayout.DontWrapRows)
        rep_layout.setContentsMargins(10, 8, 10, 6)
        rep_layout.setHorizontalSpacing(6)
        rep_layout.setVerticalSpacing(3)
        self.rep_diam_input, self.rep_offset_input = QLineEdit(), QLineEdit()
        rep_layout.addRow(self._create_form_label("Diâmetro Furos:", 92), self.rep_diam_input)
        rep_layout.addRow(self._create_form_label("Offset Borda:", 92), self.rep_offset_input)
        self.replicate_btn = QPushButton("Replicar Furos")
        self.replicate_btn.setMinimumHeight(22)
        self.replicate_btn.setMaximumHeight(24)
        rep_layout.addRow(self.replicate_btn)
        self.rep_group.setLayout(rep_layout)
        self.rep_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Maximum)
        self.rep_group.setMinimumHeight(84)
        furos_main_layout.addWidget(self.rep_group)
        man_group = QGroupBox("Furos Manuais")
        man_layout = QVBoxLayout()
        man_layout.setContentsMargins(10, 8, 10, 8)
        man_layout.setSpacing(4)
        man_form_layout = QGridLayout()
        man_form_layout.setContentsMargins(0, 0, 0, 0)
        man_form_layout.setHorizontalSpacing(6)
        man_form_layout.setVerticalSpacing(2)
        self.diametro_furo_input, self.pos_x_input, self.pos_y_input = QLineEdit(), QLineEdit(), QLineEdit()
        manual_hole_rows = [
            ("Diâmetro:", self.diametro_furo_input),
            ("Posição X:", self.pos_x_input),
            ("Posição Y:", self.pos_y_input),
        ]
        for row_index, (text, field) in enumerate(manual_hole_rows):
            field.setMinimumHeight(18)
            field.setMaximumHeight(22)
            man_form_layout.addWidget(self._create_form_label(text, 72), row_index, 0)
            man_form_layout.addWidget(field, row_index, 1)
        man_form_layout.setColumnStretch(1, 1)
        self.add_furo_btn = QPushButton("Adicionar Furo Manual")
        self.add_furo_btn.setMinimumHeight(22)
        self.add_furo_btn.setMaximumHeight(24)
        man_layout.addLayout(man_form_layout)
        man_layout.addWidget(self.add_furo_btn)
        self.furos_table = QTableWidget(0, 4)
        self.furos_table.setMinimumHeight(96)
        self.furos_table.setMaximumHeight(122)
        self.furos_table.setAlternatingRowColors(True)
        self.furos_table.setHorizontalHeaderLabels(["Diâmetro", "Pos X", "Pos Y", "Ação"])
        man_layout.addWidget(self.furos_table)
        man_group.setLayout(man_layout)
        man_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
        man_group.setMinimumHeight(182)
        furos_main_layout.addWidget(man_group)
        furos_main_group.setLayout(furos_main_layout)
        furos_main_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        list_group = QGroupBox("6. Lista de Peças para Produção")
        list_layout = QVBoxLayout()
        list_layout.setContentsMargins(10, 10, 10, 8)
        list_layout.setSpacing(6)
        self.pieces_table = QTableWidget()
        self.table_headers = [col.replace('_', ' ').title() for col in self.colunas_df] + ["Ações"]
        self.pieces_table.setColumnCount(len(self.table_headers))
        self.pieces_table.setHorizontalHeaderLabels(self.table_headers)
        self.pieces_table.verticalHeader().setDefaultSectionSize(28) 
        self.pieces_table.setMinimumHeight(156)
        self.pieces_table.setAlternatingRowColors(True)
        self.pieces_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        list_layout.addWidget(self.pieces_table, 1)
        self.dir_label = QLabel("Nenhum projeto ativo. Inicie um novo projeto.")
        self.dir_label.setStyleSheet("font-style: italic;")
        self.dir_label.setWordWrap(True)
        list_layout.addWidget(self.dir_label)
        process_buttons_layout = QGridLayout()
        process_buttons_layout.setContentsMargins(0, 0, 0, 0)
        process_buttons_layout.setHorizontalSpacing(5)
        process_buttons_layout.setVerticalSpacing(5)
        self.conclude_project_btn = QPushButton("Projeto Concluído")
        self.export_excel_btn = QPushButton("Exportar Excel")
        self.process_pdf_btn, self.process_dxf_btn, self.process_all_btn = QPushButton("Gerar PDF"), QPushButton("Gerar DXF"), QPushButton("PDF + DXF")
        self.calculate_nesting_btn = QPushButton("Aproveitamento")
        for button in [self.export_excel_btn, self.conclude_project_btn, self.calculate_nesting_btn, self.process_pdf_btn, self.process_dxf_btn, self.process_all_btn]:
            button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            button.setMinimumHeight(22)
            button.setMaximumHeight(24)
        process_buttons_layout.addWidget(self.export_excel_btn, 0, 0)
        process_buttons_layout.addWidget(self.conclude_project_btn, 0, 1)
        process_buttons_layout.addWidget(self.calculate_nesting_btn, 1, 0)
        process_buttons_layout.addWidget(self.process_pdf_btn, 1, 1)
        process_buttons_layout.addWidget(self.process_dxf_btn, 2, 0)
        process_buttons_layout.addWidget(self.process_all_btn, 2, 1)
        process_buttons_layout.setColumnStretch(0, 1)
        process_buttons_layout.setColumnStretch(1, 1)
        list_layout.addLayout(process_buttons_layout)
        list_group.setLayout(list_layout)
        list_group.setMinimumHeight(306)
        list_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        right_panel_widget = QWidget()
        right_v_layout = QVBoxLayout(right_panel_widget)
        right_v_layout.setContentsMargins(0, 0, 0, 0)
        right_v_layout.setSpacing(10)
        right_panel_widget.setMinimumWidth(470)
        right_panel_widget.setMaximumWidth(610)
        right_panel_widget.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        right_v_layout.addWidget(furos_main_group, 4)
        right_v_layout.addWidget(list_group, 5)

        top_h_layout.addWidget(left_panel_widget, stretch=14, alignment=Qt.AlignTop)
        top_h_layout.addWidget(right_panel_widget, stretch=8, alignment=Qt.AlignTop)

        top_container_widget = QWidget()
        top_container_widget.setLayout(top_h_layout)
        top_container_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)


        log_group = QGroupBox("Log de Execução")
        log_layout = QVBoxLayout()
        self.log_text = QTextEdit()
        self.log_text.setObjectName("logExecution")
        self.log_text.setReadOnly(True)
        log_layout.addWidget(self.log_text)
        log_group.setLayout(log_layout)
        log_group.setMinimumHeight(88)
        log_group.setMaximumHeight(128)

        main_layout.addWidget(top_container_widget, 1)
        main_layout.addWidget(log_group, 0)


        self.progress_bar = QProgressBar()
        main_layout.addWidget(self.progress_bar)
        
        self.statusBar().showMessage("Pronto")
        

        self.start_project_btn.setObjectName("primaryButton")
        self.conclude_project_btn.setObjectName("successButton")
        self.calculate_nesting_btn.setObjectName("warningButton")
        self.process_all_btn.setObjectName("primaryButton")

    def connect_signals(self):
        """Método para centralizar todas as conexões de sinais e slots."""
        self.support_files_btn.clicked.connect(self.show_support_files_dialog)
        self.v3_features_btn.clicked.connect(self.show_v3_features_dialog)
        self.about_btn.clicked.connect(self.show_about_dialog)
        self.help_btn.clicked.connect(self.show_help_dialog)
        self.calculate_nesting_btn.clicked.connect(self.open_nesting_dialog)
        self.start_project_btn.clicked.connect(self.start_new_project)
        self.theme_toggle_btn.clicked.connect(self.toggle_theme) 
        self.history_btn.clicked.connect(self.show_history_dialog)
        self.select_file_btn.clicked.connect(self.select_file)
        self.import_dxf_btn.clicked.connect(self.import_dxfs) 
        self.clear_excel_btn.clicked.connect(self.clear_excel_data)
        self.select_auto_image_btn.clicked.connect(self.select_auto_upload_image)
        self.import_auto_json_btn.clicked.connect(self.import_auto_json)
        self.upload_schema_btn.clicked.connect(self.show_upload_schema_dialog)
        self.generate_code_btn.clicked.connect(self.generate_piece_code)
        self.add_piece_btn.clicked.connect(self.add_manual_piece)
        self.forma_combo.currentTextChanged.connect(self.update_dimension_fields)
        self.replicate_btn.clicked.connect(self.replicate_holes)
        self.add_furo_btn.clicked.connect(self.add_furo_temp)
        self.process_pdf_btn.clicked.connect(self.start_pdf_generation)
        self.process_dxf_btn.clicked.connect(self.start_dxf_generation)
        self.process_all_btn.clicked.connect(self.start_all_generation)
        self.conclude_project_btn.clicked.connect(self.conclude_project)
        self.export_excel_btn.clicked.connect(self.export_project_to_excel)

    def toggle_theme(self):
        """(NOVA FUNÇÃO) Alterna entre o tema claro e escuro."""
        self.is_dark_theme = not self.is_dark_theme
        if self.is_dark_theme:
            self.theme_toggle_btn.setText("☀️ Tema Claro")
            QApplication.instance().setStyleSheet(DARK)
        else:
            self.theme_toggle_btn.setText("🌙 Tema Escuro")
            QApplication.instance().setStyleSheet(STYLE)
        self.log_text.append(f"Tema alterado para {'Escuro' if self.is_dark_theme else 'Claro'}.")

    def _get_dynamic_offset_and_margin(self, espessura, default_offset, default_margin):
        """Retorna o offset e a margem com base na espessura."""

        if abs(default_offset - 8.0) > 1e-5:
            return default_offset, default_margin

        if 0 < espessura <= 6.35: return 5, 10
        elif 6.35 < espessura <= 15.88: return 10, default_margin
        elif 15.88 < espessura <= 20: return 17, default_margin
        elif abs(espessura - 22.22) < 1e-5: return 20, default_margin
        elif 25.4 <= espessura <= 38: return 25, default_margin
        return default_offset, default_margin


    def start_new_project(self):
        parent_dir = QFileDialog.getExistingDirectory(self, "Selecione a Pasta Principal para o Novo Projeto")
        if not parent_dir: return
        project_name, ok = QInputDialog.getText(self, "Novo Projeto", "Digite o nome ou número do novo projeto:")
        if ok and project_name:
            project_path = os.path.join(parent_dir, project_name)
            if os.path.exists(project_path):
                reply = QMessageBox.question(self, 'Diretório Existente', f"A pasta '{project_name}' já existe.\nDeseja usá-la como o diretório do projeto ativo?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply == QMessageBox.No: return
            else:
                try: os.makedirs(project_path)
                except OSError as e: QMessageBox.critical(self, "Erro ao Criar Pasta", f"Não foi possível criar o diretório do projeto:\n{e}"); return
            self._clear_session(clear_project_number=True)
            self.project_directory = project_path
            self.projeto_input.setText(project_name)
            self.dir_label.setText(f"Projeto Ativo: {self.project_directory}")
            self.dir_label.setStyleSheet("font-style: normal;")
            self.log_text.append(f"\n--- NOVO PROJETO INICIADO: {project_name} ---")
            self.log_text.append(f"Arquivos serão salvos em: {self.project_directory}")
            self.set_initial_button_state()

    def set_initial_button_state(self):
        is_project_active = self.project_directory is not None
        has_items = not (self.excel_df.empty and self.manual_df.empty)
        self.calculate_nesting_btn.setEnabled(is_project_active and has_items)
        self.start_project_btn.setEnabled(True)
        self.history_btn.setEnabled(True)
        self.v3_features_btn.setEnabled(True)
        self.select_file_btn.setEnabled(is_project_active)
        self.import_dxf_btn.setEnabled(is_project_active)
        self.clear_excel_btn.setEnabled(is_project_active and not self.excel_df.empty)
        self.select_auto_image_btn.setEnabled(is_project_active)
        self.import_auto_json_btn.setEnabled(is_project_active)
        self.upload_schema_btn.setEnabled(True)
        self.generate_code_btn.setEnabled(is_project_active)
        self.add_piece_btn.setEnabled(is_project_active)
        self.replicate_btn.setEnabled(is_project_active)
        self.add_furo_btn.setEnabled(is_project_active)
        self.process_pdf_btn.setEnabled(is_project_active and has_items)
        self.process_dxf_btn.setEnabled(is_project_active and has_items)
        self.process_all_btn.setEnabled(is_project_active and has_items)
        self.conclude_project_btn.setEnabled(is_project_active and has_items)
        self.export_excel_btn.setEnabled(is_project_active and has_items)
        self.progress_bar.setVisible(False)

    def show_history_dialog(self):
        dialog = HistoryDialog(self.history_manager, self)
        if dialog.exec_() == QDialog.Accepted:
            loaded_pieces = dialog.loaded_project_data
            if loaded_pieces:
                project_number_loaded = loaded_pieces[0].get('project_number') if loaded_pieces and 'project_number' in loaded_pieces[0] else dialog.project_list_widget.currentItem().text()
                self.start_new_project_from_history(project_number_loaded, loaded_pieces)
    
    def start_new_project_from_history(self, project_name, pieces_data):
        parent_dir = QFileDialog.getExistingDirectory(self, f"Selecione uma pasta para o projeto '{project_name}'")
        if not parent_dir: return
        project_path = os.path.join(parent_dir, project_name)
        os.makedirs(project_path, exist_ok=True)
        self._clear_session(clear_project_number=True)
        self.project_directory = project_path
        self.projeto_input.setText(project_name)
        self.excel_df = pd.DataFrame(columns=self.colunas_df)
        self.manual_df = pd.DataFrame(pieces_data)
        self.dir_label.setText(f"Projeto Ativo: {self.project_directory}"); self.dir_label.setStyleSheet("font-style: normal;")
        self.log_text.append(f"\n--- PROJETO DO HISTÓRICO CARREGADO: {project_name} ---")
        self.update_table_display()
        self.set_initial_button_state()

    def start_pdf_generation(self): self.start_processing(generate_pdf=True, generate_dxf=False)
    def start_dxf_generation(self): self.start_processing(generate_pdf=False, generate_dxf=True)
    def start_all_generation(self): self.start_processing(generate_pdf=True, generate_dxf=True)

    def start_processing(self, generate_pdf, generate_dxf):
        if not self.project_directory:
            QMessageBox.warning(self, "Nenhum Projeto Ativo", "Inicie um novo projeto antes de gerar arquivos."); return
        project_number = self.projeto_input.text().strip()
        if not project_number:
            QMessageBox.warning(self, "Número do Projeto Ausente", "Por favor, defina um número para o projeto ativo."); return

        dfs_to_concat = [df for df in [self.excel_df, self.manual_df] if not df.empty]
        if not dfs_to_concat:
            QMessageBox.warning(self, "Aviso", "A lista de peças está vazia."); return
        combined_df = pd.concat(dfs_to_concat, ignore_index=True)

        self.set_buttons_enabled_on_process(False)
        self.progress_bar.setVisible(True); self.progress_bar.setValue(0); self.log_text.clear()
        self.process_thread = ProcessThread(combined_df.copy(), generate_pdf, generate_dxf, self.project_directory, project_number)
        self.process_thread.update_signal.connect(self.log_text.append)
        self.process_thread.progress_signal.connect(self.progress_bar.setValue)
        self.process_thread.finished_signal.connect(self.processing_finished)
        self.process_thread.start()

    def processing_finished(self, success, message):
        self.set_buttons_enabled_on_process(True); self.progress_bar.setVisible(False)
        msgBox = QMessageBox.information if success else QMessageBox.critical
        msgBox(self, "Concluído" if success else "Erro", message); self.statusBar().showMessage("Pronto")
    
    def conclude_project(self):
        project_number = self.projeto_input.text().strip()
        if not project_number:
            QMessageBox.warning(self, "Projeto sem Número", "O projeto ativo não tem um número definido.")
            return
        reply = QMessageBox.question(self, 'Concluir Projeto', f"Deseja salvar e concluir o projeto '{project_number}'?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:

            dfs_to_concat = [df for df in [self.excel_df, self.manual_df] if not df.empty]
            if dfs_to_concat:
                combined_df = pd.concat(dfs_to_concat, ignore_index=True)

                combined_df['project_number'] = project_number
                combined_df['project_number'] = project_number
                self.history_manager.save_project(project_number, combined_df)
                self.log_text.append(f"Projeto '{project_number}' salvo no histórico.")
            self._clear_session(clear_project_number=True)
            self.project_directory = None
            self.dir_label.setText("Nenhum projeto ativo. Inicie um novo projeto."); self.dir_label.setStyleSheet("font-style: italic;")
            self.set_initial_button_state()
            self.log_text.append(f"\n--- PROJETO '{project_number}' CONCLUÍDO ---")

    def open_nesting_dialog(self):

        dfs_to_concat = [df for df in [self.excel_df, self.manual_df] if not df.empty]
        if not dfs_to_concat:
            QMessageBox.warning(self, "Lista Vazia", "Não há peças na lista para calcular o aproveitamento.")
            return
        combined_df = pd.concat(dfs_to_concat, ignore_index=True)

        valid_df = combined_df[combined_df['forma'].isin(['rectangle', 'circle', 'right_triangle', 'trapezoid', 'dxf_shape'])].copy()
        if valid_df.empty:
            QMessageBox.information(self, "Nenhuma Peça Válida", "O cálculo de aproveitamento só pode ser feito com peças da forma 'rectangle', 'circle', 'right_triangle', 'trapezoid' ou 'dxf_shape'.")
            return

        dialog = NestingDialog(valid_df, self)
        dialog.exec_()

    def _get_export_parameters(self):
        """Abre um diálogo para obter os parâmetros de exportação."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Parâmetros de Exportação e Aproveitamento")
        layout = QFormLayout(dialog)

        chapa_largura_input = QLineEdit("3000")
        chapa_altura_input = QLineEdit("1500")
        offset_input = QLineEdit("8")
        margin_input = QLineEdit("10")
        method_combo = QComboBox()
        method_combo.addItems(["Plasma/Laser", "Guilhotina"])

        layout.addRow("Largura da Chapa (mm):", chapa_largura_input)
        layout.addRow("Altura da Chapa (mm):", chapa_altura_input)
        layout.addRow("Método de Corte:", method_combo)
        layout.addRow("Offset entre Peças (mm) [Plasma/Laser]:", offset_input)
        layout.addRow("Margem da Chapa (mm) [Plasma/Laser]:", margin_input)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addRow(button_box)

        if dialog.exec_() == QDialog.Accepted:
            try:
                params = {
                    "chapa_largura": float(chapa_largura_input.text()),
                    "chapa_altura": float(chapa_altura_input.text()),
                    "offset": float(offset_input.text()),
                    "margin": float(margin_input.text()),
                    "method": method_combo.currentText()
                }
                return params
            except (ValueError, TypeError):
                QMessageBox.critical(self, "Erro de Entrada", "Valores de chapa, offset e margem devem ser numéricos.")
                return None
        return None

    def export_project_to_excel(self):
        params = self._get_export_parameters()
        if not params: return

        project_number = self.projeto_input.text().strip()
        if not project_number:
            QMessageBox.warning(self, "Nenhum Projeto Ativo", "Inicie um novo projeto para poder exportá-lo.")
            return

        dfs_to_concat = [df for df in [self.excel_df, self.manual_df] if not df.empty]
        if not dfs_to_concat:
            QMessageBox.warning(self, "Lista Vazia", "Não há peças na lista para exportar.")
            return
        combined_df = pd.concat(dfs_to_concat, ignore_index=True)

        default_filename = os.path.join(self.project_directory, f"Relacao-de-peças-projeto_{project_number}.xlsx")
        save_path, _ = QFileDialog.getSaveFileName(self, "Salvar Resumo do Projeto", default_filename, "Excel Files (*.xlsx)")
        if not save_path:
            return

        start_time = time.time()
        os.environ['CURRENT_PROJECT_NAME'] = project_number

        self.set_buttons_enabled_on_process(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.log_text.clear()
        self.log_text.append("Iniciando exportação para Excel...")
        QApplication.processEvents()

        try:
            template_path = find_resource_path("planilha-dbx.xlsx")
            if template_path is not None:
                wb = load_workbook(str(template_path))
                ws = wb.active
            else:
                self.log_text.append("Template 'planilha-dbx.xlsx' nao encontrado. Criando nova planilha estruturada.")
                wb = Workbook()
                ws = wb.active
                ws.title = "PEÇAS PADRÃO"
                
                headers = [
                    "ID do Cliente", "ID Noroaco", "QTD de Pecas", "TIPO PEÇA", "Nº FUROS", "DIAMETRO FURO", 
                    "ESPESSURA", "DIMENSÃO (A)mm", "DIMENSÃO (B)mm", "DIMENSÃO (C)mm", 
                    " MATERIA PRIMA LARGURA CHAPA mm", "Kg", "Kg Total", "Tempo", "Preço por Peça (Individual)", 
                    "Total", "MÁQ", "MARGEM", "REGRA", "Sobra?", "Valor Sugerido (KG)", 
                    "Calculo Considerando Margem 30%", "Calculo sem Considerar Margem 30%", "multiplicacao CUSTO", 
                    "Prod por Min", "Forma de Desenho", "Custo Máq", "PREÇO M.P", "Imposto", "Comissão", 
                    "Frete", "Perda", "Margem", "Veloc teorica", "Fator chapa / sucata", "Tempo corte min", 
                    "Fator furo (min)", "Fator peças"
                ]
                
                ws['A1'] = "Imposto (%)"
                ws['C1'] = "Frete (R$)"
                
                for i, header in enumerate(headers, 1):
                    ws.cell(row=3, column=i, value=header).font = Font(bold=True)

            try:
                imposto_val = float(self.imposto_input.text().replace(',', '.'))
                frete_val = float(self.frete_input.text().replace(',', '.'))
                ws['A2'] = imposto_val
                ws['C2'] = frete_val
                self.log_text.append(f"Imposto ({imposto_val}) e Frete ({frete_val}) preenchidos nas células A2 e C2.")
            except ValueError:
                QMessageBox.warning(self, "Valor Inválido", "Valores de Imposto e Frete devem ser numéricos. Usando 0.")
                ws['A2'] = 0
                ws['C2'] = 0

            self.log_text.append("Preenchendo lista de peças...")
            QApplication.processEvents()
            
            start_row = 4 
            last_filled_row = start_row - 1 


            todas_as_sobras_aproveitaveis = []

            # OTIMIZACAO: Converter para lista de dicionários é muito mais rápido que iterrows()
            records = combined_df.to_dict('records')
            total_records = len(records)
            
            for index, row_data in enumerate(records):
                current_row = start_row + index
                last_filled_row = current_row 
                
                ws.cell(row=current_row, column=1, value=project_number)
                ws.cell(row=current_row, column=2, value=row_data.get('nome_arquivo', ''))
                
                qtd_peca = row_data.get('qtd', 0)
                ws.cell(row=current_row, column=3, value=qtd_peca)
                
                forma = str(row_data.get('forma', '')).lower()
                largura, altura = row_data.get('largura', 0), row_data.get('altura', 0)
                forma_map = {'circle': 'C', 'trapezoid': 'TP', 'right_triangle': 'T'}
                forma_abreviada = 'Q' if forma == 'rectangle' and largura == altura and largura > 0 else forma_map.get(forma, 'R' if forma == 'rectangle' else '')
                ws.cell(row=current_row, column=4, value=forma_abreviada)

                furos = row_data.get('furos', [])
                num_furos = len(furos) if isinstance(furos, list) else 0
                ws.cell(row=current_row, column=5, value=num_furos)
                ws.cell(row=current_row, column=6, value=furos[0].get('diam', 0) if num_furos > 0 else 0)
                

                espessura_peca = row_data.get('espessura', 0)
                ws.cell(row=current_row, column=7, value=espessura_peca) 
                
                ws.cell(row=current_row, column=8, value=largura)
                ws.cell(row=current_row, column=9, value=altura)
                
                # OTIMIZACAO: Atualizar UI apenas a cada 10 itens para não travar o processamento
                if index % 10 == 0 or index == total_records - 1:
                    self.progress_bar.setValue(int(((index + 1) / (total_records * 2)) * 100))
                    QApplication.processEvents()

            self.log_text.append("Calculando aproveitamento de chapas...")
            QApplication.processEvents()

            valid_nesting_df = combined_df[combined_df['forma'].isin(['rectangle', 'circle', 'right_triangle', 'trapezoid', 'dxf_shape'])].copy()
            valid_nesting_df['espessura'] = valid_nesting_df['espessura'].astype(float)
            

            grouped = valid_nesting_df.groupby('espessura')
            
            current_row = 212 
            ws.cell(row=current_row, column=1, value="RELATÓRIO DE APROVEITAMENTO DE CHAPA").font = Font(bold=True, size=14)
            current_row += 2


            total_perca_ponderada_real = 0.0
            total_pecas_contadas_real = 0.0
            perda_results_map = {}


            for espessura, group in grouped:
                is_guillotine = params["method"] == "Guilhotina"
                
                if is_guillotine:

                    current_offset, refila = 0, 2 * espessura
                    sheet_width_for_calc, sheet_height_for_calc = params["chapa_largura"] - refila, params["chapa_altura"]
                    effective_margin = 0
                else: 

                    current_offset, _ = self._get_dynamic_offset_and_margin(espessura, params["offset"], params["margin"])
                    effective_margin = 10 - (current_offset / 2)
                    sheet_width_for_calc, sheet_height_for_calc = params["chapa_largura"], params["chapa_altura"]

                pecas_para_calcular = []
                total_pecas_neste_grupo = 0

                for _, row in group.iterrows():
                    qtd = int(row['qtd'])
                    total_pecas_neste_grupo += qtd

                    if row['forma'] == 'rectangle' and row['largura'] > 0 and row['altura'] > 0:
                        pecas_para_calcular.append({'forma': 'rectangle', 'largura': row['largura'], 'altura': row['altura'], 'quantidade': qtd})
                    elif row['forma'] == 'circle' and row['diametro'] > 0:
                        pecas_para_calcular.append({'forma': 'circle', 'largura': row['diametro'], 'altura': row['diametro'], 'diametro': row['diametro'], 'quantidade': qtd})
                    elif row['forma'] == 'right_triangle' and row['rt_base'] > 0 and row['rt_height'] > 0:
                        pecas_para_calcular.append({'forma': 'right_triangle', 'largura': row['rt_base'], 'altura': row['rt_height'], 'quantidade': qtd})
                    elif row['forma'] == 'trapezoid' and row['trapezoid_large_base'] > 0 and row['trapezoid_height'] > 0:
                        pecas_para_calcular.append({'forma': 'trapezoid', 'largura': row['trapezoid_large_base'], 'altura': row['trapezoid_height'], 'small_base': row['trapezoid_small_base'], 'quantidade': qtd})
                    elif row['forma'] == 'dxf_shape' and row['largura'] > 0 and row['altura'] > 0:
                        pecas_para_calcular.append({'forma': 'dxf_shape', 'largura': row['largura'], 'altura': row['altura'], 'dxf_path': row['dxf_path'], 'quantidade': qtd})

                if not pecas_para_calcular: continue

                
                pecas_com_offset = []
                for p in pecas_para_calcular:
                    p_copy = p.copy()
                    p_copy['largura'] += current_offset
                    p_copy['altura'] += current_offset
                    if 'small_base' in p_copy: p_copy['small_base'] += current_offset
                    pecas_com_offset.append(p_copy)
                
                # OTIMIZACAO: Log menos frequente
                self.log_text.append(f"Otimizando espessura {espessura}mm...")
                QApplication.processEvents()

                resultado = orquestrar_planos_de_corte(sheet_width_for_calc, sheet_height_for_calc, pecas_com_offset, current_offset, effective_margin, espessura, is_guillotine, status_signal_emitter=None)
                
                if not resultado: continue


                for plano in resultado.get('planos_unicos', []):
                    for sobra in plano.get('sobras', []):
                        if sobra.get('tipo_sobra') == 'aproveitavel':

                            sobra['espessura'] = espessura
                            sobra['qtd'] = plano.get('repeticoes', 1)
                            todas_as_sobras_aproveitaveis.append(sobra)

                percentual_perda = resultado.get('percentual_perda_total_sucata', 0)
                

                perda_results_map[espessura] = percentual_perda 
                
                total_perca_ponderada_real += (percentual_perda * total_pecas_neste_grupo)
                total_pecas_contadas_real += total_pecas_neste_grupo


                ws.cell(row=current_row, column=1, value=f"Espessura: {espessura} mm").font = Font(bold=True, size=12)
                current_row += 1
                total_chapas_usadas = resultado['total_chapas']
                peso_total_chapas_kg = (params["chapa_largura"]/1000) * (params["chapa_altura"]/1000) * espessura * 7.85 * total_chapas_usadas
                ws.cell(row=current_row, column=1, value=f"Total de Chapas: {total_chapas_usadas}")
                ws.cell(row=current_row, column=2, value=f"Aproveitamento: {resultado['aproveitamento_geral']}")
                ws.cell(row=current_row, column=3, value=f"Peso Total das Chapas: {peso_total_chapas_kg:.2f} kg").font = Font(bold=True)
                current_row += 2

                for i, plano_info in enumerate(resultado['planos_unicos']):

                    ws.cell(row=current_row, column=1, value=f"Plano de Corte {i+1} (Repetir {plano_info['repeticoes']}x)").font = Font(italic=True)
                    current_row += 1
                    ws.cell(row=current_row, column=2, value="Peças neste plano:")
                    current_row += 1
                    for item in plano_info['resumo_pecas']:
                        ws.cell(row=current_row, column=3, value=f"- {item['qtd']}x de {item['tipo']}")
                        current_row += 1
                    current_row += 1

                sucata_info = resultado.get('sucata_detalhada')
                if sucata_info:

                    bold_font = Font(bold=True)
                    ws.cell(row=current_row, column=1, value="Peso do Offset (perda de corte):").font = bold_font
                    ws.cell(row=current_row, column=2, value=f"{sucata_info['peso_offset']:.2f} kg")
                    current_row += 2
                    ws.cell(row=current_row, column=1, value="Sobras Aproveitáveis (Retalhos > 300x300 mm)").font = bold_font
                    current_row += 1
                    sobras_aproveitaveis = sucata_info['sobras_aproveitaveis']
                    if not sobras_aproveitaveis:
                        ws.cell(row=current_row, column=2, value="- Nenhuma")
                        current_row += 1
                    else:
                        from collections import Counter
                        contagem = Counter((s['largura'], s['altura'], f"{s['peso']:.2f}") for s in sobras_aproveitaveis for _ in range(s['quantidade']))
                        total_peso_aproveitavel = sum(s['peso'] * s['quantidade'] for s in sobras_aproveitaveis)
                        for (larg, alt, peso_unit), qtd in contagem.items():
                            ws.cell(row=current_row, column=2, value=f"- {qtd}x de {larg:.0f}x{alt:.0f} mm (Peso unit: {peso_unit} kg)")
                            current_row += 1
                        ws.cell(row=current_row, column=2, value=f"Peso Total Aproveitável: {total_peso_aproveitavel:.2f} kg").font = bold_font
                        current_row += 1
                    current_row += 1
                    ws.cell(row=current_row, column=1, value="Sucatas com Dimensões").font = bold_font
                    current_row += 1
                    sucatas_dim = sucata_info['sucatas_dimensionadas']
                    if not sucatas_dim:
                        ws.cell(row=current_row, column=2, value="- Nenhuma")
                        current_row += 1
                    else:
                        from collections import Counter
                        contagem = Counter((s['largura'], s['altura'], f"{s['peso']:.2f}") for s in sucatas_dim for _ in range(s['quantidade']))
                        total_peso_sucata_dim = sum(s['peso'] * s['quantidade'] for s in sucatas_dim)
                        for (larg, alt, peso_unit), qtd in contagem.items():
                            ws.cell(row=current_row, column=2, value=f"- {qtd}x de {larg:.0f}x{alt:.0f} mm (Peso unit: {peso_unit} kg)")
                            current_row += 1
                        ws.cell(row=current_row, column=2, value=f"Peso Total (Sucata Dimensionada): {total_peso_sucata_dim:.2f} kg").font = bold_font
                        current_row += 1
                    current_row += 1
                    ws.cell(row=current_row, column=1, value="Demais Sucatas (cavacos, etc):").font = bold_font
                    ws.cell(row=current_row, column=2, value=f"{sucata_info['peso_demais_sucatas']:.2f} kg")
                    current_row += 2
                    ws.cell(row=current_row, column=1, value="Resumo da Perda Total (Sucata + Processo + Offset):").font = bold_font
                    ws.cell(row=current_row, column=2, value=f"{resultado.get('peso_perda_total_sucata', 0):.2f} kg")
                    ws.cell(row=current_row, column=3, value=f"({resultado.get('percentual_perda_total_sucata', 0):.2f} % do total)").font = Font(italic=True)
                    current_row += 2

                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
                cell = ws.cell(row=current_row, column=1)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                current_row += 2
                
                # Atualiza barra de progresso baseada no progresso geral estimado
                self.progress_bar.setValue(50 + int(50 * (list(grouped.groups.keys()).index(espessura) + 1) / len(grouped)))

            project_name_upper = project_number.upper()
            is_special_material = any(keyword in project_name_upper for keyword in ['FF', 'GALV', 'XADREZ'])

            if is_special_material and todas_as_sobras_aproveitaveis:
                self.log_text.append("Material especial detectado. Adicionando sobras aproveitáveis à lista de peças...")
                QApplication.processEvents()

                sobras_agrupadas = {}
                for s in todas_as_sobras_aproveitaveis:
                    chave = (round(s['largura']), round(s['altura']), s['espessura'])
                    if chave not in sobras_agrupadas:
                        sobras_agrupadas[chave] = {'qtd': 0, 'largura': s['largura'], 'altura': s['altura'], 'espessura': s['espessura']}
                    sobras_agrupadas[chave]['qtd'] += s['qtd']


                for sobra_agrupada in sobras_agrupadas.values():
                    last_filled_row += 1
                    
                    ws.cell(row=last_filled_row, column=1, value=project_number)
                    ws.cell(row=last_filled_row, column=2, value="SOBRA") # Nome da peça
                    ws.cell(row=last_filled_row, column=3, value=sobra_agrupada['qtd']) # Quantidade
                    ws.cell(row=last_filled_row, column=4, value='R') # Forma (Retângulo)
                    ws.cell(row=last_filled_row, column=5, value=0) # Furos
                    ws.cell(row=last_filled_row, column=6, value=0) # Diâmetro Furo
                    ws.cell(row=last_filled_row, column=7, value=sobra_agrupada['espessura']) # Espessura
                    ws.cell(row=last_filled_row, column=8, value=sobra_agrupada['largura']) # Largura
                    ws.cell(row=last_filled_row, column=9, value=sobra_agrupada['altura']) # Altura

                start_hide_row = last_filled_row + 1
                end_hide_row = 207 
                if start_hide_row <= end_hide_row:
                    ws.row_dimensions.group(start_hide_row, end_hide_row, hidden=True)
                    self.log_text.append(f"Linhas de {start_hide_row} a {end_hide_row} re-ocultadas.")
            else:
                if not is_special_material:
                    self.log_text.append("Projeto não é de material especial. Sobras não serão adicionadas à planilha.")


            if total_pecas_contadas_real > 0:
                avg_loss_real = total_perca_ponderada_real / total_pecas_contadas_real if total_pecas_contadas_real > 0 else 0

                ws['D2'] = avg_loss_real / 100.0 
                self.log_text.append(f"Perca média ponderada REAL ({avg_loss_real:.2f}%) preenchida em D2.")
            else:
                ws['D2'] = 0
                self.log_text.append("Nenhuma peça para calcular perca real. Preenchido 0 em D2.")

            self.log_text.append("Atualizando tabela de perdas (Coluna W) com resultados do nesting...")

            perda_map_arredondado = {round(float(k), 2): v for k, v in perda_results_map.items()}
      
            start_row = 213
            num_rows = 25
            end_row_exclusive = start_row + num_rows 
            
            self.log_text.append(f"Atualizando {num_rows} linhas da tabela de perdas (V{start_row}:W{end_row_exclusive - 1})...")

 
            for row_idx in range(start_row, end_row_exclusive):  

                esp_cell = ws.cell(row=row_idx, column=22) 
                

                if esp_cell.value is None or str(esp_cell.value).strip() == "": 
                    ws.cell(row=row_idx, column=23, value=None) # Coluna W
                    continue
                    
                try:

                    
                    esp_valor_str = str(esp_cell.value).replace(',', '.')
                    esp_template = round(float(esp_valor_str), 2)
                    

                    if esp_template in perda_map_arredondado:

                        perda_para_escrever = perda_map_arredondado[esp_template] / 100.0
                        ws.cell(row=row_idx, column=23, value=perda_para_escrever)
                    else:

                        ws.cell(row=row_idx, column=23, value=0.0)
                        
                except (ValueError, TypeError):

                    self.log_text.append(f"AVISO: Valor não numérico na célula V{row_idx}: '{esp_cell.value}'. Deixando em branco.")
                    ws.cell(row=row_idx, column=23, value=None)
                    continue


            try:

                start_hide_row = last_filled_row + 1
                end_hide_row = 207 
                
                if start_hide_row <= end_hide_row:

                    ws.row_dimensions.group(start_hide_row, end_hide_row, hidden=True)
                    self.log_text.append(f"Linhas da {start_hide_row} até {end_hide_row} ocultadas com sucesso.")
                else:

                    self.log_text.append(f"Nenhuma linha para ocultar (Última linha preenchida: {last_filled_row}).")
            except Exception as e:
                self.log_text.append(f"AVISO: Falha ao ocultar linhas. {e}")

            self.log_text.append("Salvando arquivo Excel...")
            QApplication.processEvents()
            wb.save(save_path)
            self.progress_bar.setValue(100)
            self.log_text.append(f"Resumo do projeto salvo com sucesso em: {save_path}")
            QMessageBox.information(self, "Sucesso", f"O arquivo Excel foi salvo com sucesso em:\n{save_path}")
            #self._generate_pdf_from_excel(save_path, len(combined_df))
        except Exception as e:
            self.log_text.append(f"ERRO ao exportar para Excel: {e}")
            QMessageBox.critical(self, "Erro na Exportação", f"Ocorreu um erro ao salvar o arquivo:\n{e}")
        finally:
            elapsed_time = time.time() - start_time
            self.log_text.append(f"Tempo de execução da exportação: {elapsed_time:.2f}s")
            self.set_buttons_enabled_on_process(True)
            self.progress_bar.setVisible(False)


            if 'CURRENT_PROJECT_NAME' in os.environ:
                del os.environ['CURRENT_PROJECT_NAME']

        

    # def _generate_pdf_from_excel(self, excel_path, num_pecas):
    #     """
    #     (NOVA FUNÇÃO)
    #     Usa pywin32 para abrir o Excel salvo e exportá-lo como PDF.
    #     Esta lógica foi portada do seu script CLI.
    #     """
    #     if not PYWIN32_DISPONIVEL:
    #         self.log_text.append("\n[AVISO] Geração de PDF pulada. Biblioteca 'pywin32' não encontrada.")
    #         return

    #     self.log_text.append("\nIniciando geração de PDF do orçamento...")
        
    #     pdf_filename = os.path.splitext(excel_path)[0] + ".pdf"
    #     HEADER_ROW = 3 
    #     START_ROW = 4 
    #     TOTAL_ROW = 209 
    #     LAST_EMPTY_ROW = 207 

    #     excel = None
    #     workbook = None

    #     try:
    #         last_data_row = START_ROW + num_pecas - 1
    #         first_empty_row = last_data_row + 1
            
    #         range_to_hide = None
    #         if first_empty_row <= LAST_EMPTY_ROW:
    #             range_to_hide = f"{first_empty_row}:{LAST_EMPTY_ROW}"
            
    #         pythoncom.CoInitialize()
    #         excel = win32com.client.Dispatch("Excel.Application")
    #         excel.Visible = False
    #         excel.DisplayAlerts = False
            
    #         full_excel_path = os.path.abspath(excel_path)
            
    #         workbook = excel.Workbooks.Open(full_excel_path)
            

    #         sheet = workbook.Worksheets("PEÇAS PADRÃO")
    #         sheet.Activate()
            
    #         sheet.Rows.Hidden = False 
    #         if range_to_hide:
    #             self.log_text.append(f"Ocultando linhas {range_to_hide} para o PDF...")
    #             sheet.Rows(range_to_hide).Hidden = True
            

    #         print_area_range = f"A{HEADER_ROW}:V{TOTAL_ROW}" 
    #         sheet.PageSetup.PrintArea = print_area_range
            
    #         sheet.PageSetup.Zoom = False
    #         sheet.PageSetup.FitToPagesWide = 1 
    #         sheet.PageSetup.FitToPagesTall = 1 
    #         sheet.PageSetup.Orientation = 2
            
    #         full_pdf_path = os.path.abspath(pdf_filename)
    #         self.log_text.append(f"Exportando PDF para: {full_pdf_path}...")
            

    #         sheet.ExportAsFixedFormat(0, full_pdf_path)
            
    #         self.log_text.append(f"✅ SUCESSO! PDF do orçamento gerado.")

    #     except Exception as e:
    #         self.log_text.append(f"\n[ERRO] Falha ao gerar o PDF do orçamento: {e}")
    #         self.log_text.append("Verifique se o Excel está instalado e se o pywin32 foi registrado (pywin32_postinstall.py -install).")
        
    #     finally:

    #         try:
    #             if workbook:
    #                 workbook.Close(SaveChanges=False) 
    #             if excel:
    #                 excel.Quit()
    #             del excel 
    #             pythoncom.CoUninitialize()
    #         except Exception as e_cleanup:
    #             self.log_text.append(f"[AVISO] Erro durante a limpeza do COM: {e_cleanup}")
    #             try:
                    
    #                 pythoncom.CoUninitialize()
    #             except:
    #                 pass 

    def _clear_session(self, clear_project_number=False):
        fields_to_clear = [self.nome_input, self.espessura_input, self.qtd_input, self.largura_input, self.altura_input, self.diametro_input, self.rt_base_input, self.rt_height_input, self.trapezoid_large_base_input, self.trapezoid_small_base_input, self.trapezoid_height_input, self.rep_diam_input, self.rep_offset_input, self.diametro_furo_input, self.pos_x_input, self.pos_y_input]
        if clear_project_number:
            fields_to_clear.append(self.projeto_input)
        for field in fields_to_clear:
            field.clear()
        self.furos_atuais = []
        self.auto_upload_image_path = None
        self.auto_upload_json_path = None
        self.upload_status_label.setText("Nenhuma base visual ou JSON automático selecionado.")
        self.update_furos_table()
        self.file_label.setText("Nenhum projeto ativo.")
        if clear_project_number: 
            self.excel_df = pd.DataFrame(columns=self.colunas_df)
            self.manual_df = pd.DataFrame(columns=self.colunas_df)
            self.update_table_display()

    def set_buttons_enabled_on_process(self, enabled):
        is_project_active = self.project_directory is not None
        has_items = not (self.excel_df.empty and self.manual_df.empty)
        self.calculate_nesting_btn.setEnabled(enabled and is_project_active and has_items)
        self.start_project_btn.setEnabled(enabled)
        self.history_btn.setEnabled(enabled)
        self.theme_toggle_btn.setEnabled(enabled)
        self.v3_features_btn.setEnabled(enabled)
        self.select_file_btn.setEnabled(enabled and is_project_active)
        self.import_dxf_btn.setEnabled(enabled and is_project_active) 
        self.clear_excel_btn.setEnabled(enabled and is_project_active and not self.excel_df.empty)
        self.select_auto_image_btn.setEnabled(enabled and is_project_active)
        self.import_auto_json_btn.setEnabled(enabled and is_project_active)
        self.upload_schema_btn.setEnabled(enabled)
        self.generate_code_btn.setEnabled(enabled and is_project_active)
        self.add_piece_btn.setEnabled(enabled and is_project_active)
        self.replicate_btn.setEnabled(enabled and is_project_active)
        self.add_furo_btn.setEnabled(enabled and is_project_active)
        self.process_pdf_btn.setEnabled(enabled and is_project_active and has_items)
        self.process_dxf_btn.setEnabled(enabled and is_project_active and has_items)
        self.process_all_btn.setEnabled(enabled and is_project_active and has_items)
        self.conclude_project_btn.setEnabled(enabled and is_project_active and has_items)
        self.export_excel_btn.setEnabled(enabled and is_project_active and has_items)

    def update_table_display(self):
        self.set_initial_button_state()
        
        dfs_to_concat = [df for df in [self.excel_df, self.manual_df] if not df.empty]
        if dfs_to_concat:
            combined_df = pd.concat(dfs_to_concat, ignore_index=True)
        else:
            combined_df = pd.DataFrame(columns=self.colunas_df)
        self.pieces_table.blockSignals(True)
        self.pieces_table.setRowCount(0)
        self.pieces_table.blockSignals(False)

        if combined_df.empty:
            return

        self.pieces_table.setRowCount(len(combined_df))
        self.pieces_table.verticalHeader().setDefaultSectionSize(40)
        
        for i, row in combined_df.iterrows():
            for j, col in enumerate(self.colunas_df):
                value = row.get(col)
                if col == 'furos' and isinstance(value, list):
                    display_value = f"{len(value)} Furo(s)"
                elif pd.isna(value) or value == 0:
                    display_value = '-'
                else:
                    display_value = str(value)
                item = QTableWidgetItem(display_value)
                item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                self.pieces_table.setItem(i, j, item)

            action_widget = QWidget()
            action_layout = QHBoxLayout(action_widget)
            action_layout.setContentsMargins(5, 0, 5, 0)
            action_layout.setSpacing(5)
            edit_btn, delete_btn = QPushButton("Editar"), QPushButton("Excluir")
            edit_btn.clicked.connect(lambda _, r=i: self.edit_row(r))
            delete_btn.clicked.connect(lambda _, r=i: self.delete_row(r))
            action_layout.addWidget(edit_btn)
            action_layout.addWidget(delete_btn)
            self.pieces_table.setCellWidget(i, len(self.colunas_df), action_widget)

        header = self.pieces_table.horizontalHeader()
        header_map = {self.table_headers[i]: i for i in range(len(self.table_headers))}

        for col_name in ['Forma', 'Espessura', 'Qtd', 'Furos']:
            if col_name in header_map:
                header.setSectionResizeMode(header_map[col_name], QHeaderView.ResizeToContents)
        
        if 'Nome Arquivo' in header_map:
            header.setSectionResizeMode(header_map['Nome Arquivo'], QHeaderView.Stretch)
            
        dim_cols = ['Largura', 'Altura', 'Diametro', 'Rt Base', 'Rt Height', 
                    'Trapezoid Large Base', 'Trapezoid Small Base', 'Trapezoid Height']
        for col_name in dim_cols:
            if col_name in header_map:
                header.setSectionResizeMode(header_map[col_name], QHeaderView.ResizeToContents)

        if 'Ações' in header_map:
            header.setSectionResizeMode(header_map['Ações'], QHeaderView.ResizeToContents)

    def edit_row(self, row_index):
        len_excel = len(self.excel_df)
        is_from_excel = row_index < len_excel
        df_source = self.excel_df if is_from_excel else self.manual_df
        local_index = row_index if is_from_excel else row_index - len_excel
        if local_index >= len(df_source): return 
        piece_data = df_source.iloc[local_index]
        self.nome_input.setText(str(piece_data.get('nome_arquivo', '')))
        self.espessura_input.setText(str(piece_data.get('espessura', '')))
        self.qtd_input.setText(str(piece_data.get('qtd', '')))
        shape = piece_data.get('forma', '')
        index = self.forma_combo.findText(shape, Qt.MatchFixedString)
        if index >= 0: self.forma_combo.setCurrentIndex(index)
        self.largura_input.setText(str(piece_data.get('largura', '')))
        self.altura_input.setText(str(piece_data.get('altura', '')))
        self.diametro_input.setText(str(piece_data.get('diametro', '')))
        self.rt_base_input.setText(str(piece_data.get('rt_base', '')))
        self.rt_height_input.setText(str(piece_data.get('rt_height', '')))
        self.trapezoid_large_base_input.setText(str(piece_data.get('trapezoid_large_base', '')))
        self.trapezoid_small_base_input.setText(str(piece_data.get('trapezoid_small_base', '')))
        self.trapezoid_height_input.setText(str(piece_data.get('trapezoid_height', '')))
        self.furos_atuais = piece_data.get('furos', []).copy() if isinstance(piece_data.get('furos'), list) else []
        self.update_furos_table()
        df_source.drop(df_source.index[local_index], inplace=True)
        df_source.reset_index(drop=True, inplace=True)
        self.log_text.append(f"Peça '{piece_data['nome_arquivo']}' carregada para edição.")
        self.update_table_display()
    
    def delete_row(self, row_index):
        len_excel = len(self.excel_df)
        is_from_excel = row_index < len_excel
        df_source = self.excel_df if is_from_excel else self.manual_df
        local_index = row_index if is_from_excel else row_index - len_excel
        if local_index >= len(df_source): return 
        piece_name = df_source.iloc[local_index]['nome_arquivo']
        df_source.drop(df_source.index[local_index], inplace=True)
        df_source.reset_index(drop=True, inplace=True)
        self.log_text.append(f"Peça '{piece_name}' removida.")
        self.update_table_display()
    
    def generate_piece_code(self):
        """Gera um novo código para a peça com base no número do projeto. SESSÃO: CODIGO_GENERATOR"""
        project_number = self.projeto_input.text().strip()
        if not project_number: QMessageBox.warning(self, "Campo Obrigatório", "Inicie um projeto para definir o 'Nº do Projeto'."); return
        new_code = self.code_generator.generate_new_code(project_number)
        if new_code:
            self.nome_input.setText(new_code)
            self.log_text.append(f"Código '{new_code}' gerado para o projeto '{project_number}'.")
    
    def add_manual_piece(self):
        try:
            nome = self.nome_input.text().strip()
            if not nome: 
                QMessageBox.warning(self, "Campo Obrigatório", "'Nome/ID da Peça' é obrigatório.")
                return

            new_piece = {'furos': self.furos_atuais.copy()}

            for col in self.colunas_df:
                if col not in new_piece:
                    new_piece[col] = 0.0 


            new_piece.update({
                'nome_arquivo': nome, 
                'forma': self.forma_combo.currentText()
            })
            
            fields_map = { 
                'espessura': self.espessura_input, 
                'qtd': self.qtd_input, 
                'largura': self.largura_input, 
                'altura': self.altura_input, 
                'diametro': self.diametro_input, 
                'rt_base': self.rt_base_input, 
                'rt_height': self.rt_height_input, 
                'trapezoid_large_base': self.trapezoid_large_base_input, 
                'trapezoid_small_base': self.trapezoid_small_base_input, 
                'trapezoid_height': self.trapezoid_height_input 
            }
            
            for key, field in fields_map.items():
                new_piece[key] = float(field.text().replace(',', '.')) if field.text() else 0.0

            
            self.manual_df.loc[len(self.manual_df)] = new_piece

            self.log_text.append(f"Peça '{nome}' adicionada/atualizada.")
            self._clear_session(clear_project_number=False)
            self.update_table_display()
            
        except ValueError: 
            QMessageBox.critical(self, "Erro de Valor", "Campos numéricos devem conter números válidos.")
        except Exception as e:
            QMessageBox.critical(self, "Erro Inesperado", f"Ocorreu um erro ao adicionar a peça: {e}")
    
    def select_file(self):
        if not self.project_directory:
            QMessageBox.warning(self, "Nenhum Projeto Ativo", "Inicie um projeto antes de carregar uma planilha.")
            return
        
        file_path, _ = QFileDialog.getOpenFileName(self, "Selecionar Planilha", "", "Excel Files (*.xlsx *.xls)")
        if not file_path:
            return

        try:
            df = pd.read_excel(file_path, header=0, decimal=',')
            df.columns = df.columns.str.strip().str.lower()
            self.log_text.append(f"Lendo arquivo: {os.path.basename(file_path)}")

          
            rename_map = {}
            for col in df.columns:
                if 'furo_' in col and col.endswith('_diam'):
                    rename_map[col] = col.replace('_diam', '_diametro')
            if rename_map:
                df = df.rename(columns=rename_map)
                self.log_text.append(f"Colunas de diâmetro padronizadas.")

            
            max_furos = 8 
            
            furo_grupos = []
            for i in range(1, max_furos + 1):
                furo_grupos.append({
                    'diam': f'furo_{i}_diametro',
                    'x': f'furo_{i}_x',
                    'y': f'furo_{i}_y'
                })

            
            def processar_furos_da_linha(row):
                furos_encontrados = []
                for grupo in furo_grupos:
                    col_diam = grupo['diam']
                    col_x = grupo['x']
                    col_y = grupo['y']
                    
                    
                    if col_diam in row and col_x in row and col_y in row:
                        try:
                            
                            diam = pd.to_numeric(row[col_diam], errors='coerce')
                            x = pd.to_numeric(row[col_x], errors='coerce')
                            y = pd.to_numeric(row[col_y], errors='coerce')
                            
                           
                            if pd.notna(diam) and diam > 0 and pd.notna(x) and pd.notna(y):
                                furos_encontrados.append({
                                    'diam': float(diam),
                                    'x': float(x),
                                    'y': float(y)
                                })
                        except Exception:
                           
                            pass
                return furos_encontrados

            df['furos'] = df.apply(processar_furos_da_linha, axis=1)
              
            df = df.loc[:, ~df.columns.duplicated()] 
            
            
            for col in self.colunas_df:
                if col not in df.columns: 
                    df[col] = pd.NA
            
            
            numeric_cols = [
                'espessura', 'qtd', 'largura', 'altura', 'diametro', 'rt_base', 'rt_height', 
                'trapezoid_large_base', 'trapezoid_small_base', 'trapezoid_height'
            ]
            
            for col in numeric_cols:
                if col in df.columns:
                    
                    if col != 'furos':
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            
            
            self.excel_df = df[self.colunas_df].copy()
            
            self.file_label.setText(f"Planilha: {os.path.basename(file_path)}")
            self.update_table_display()
            self.log_text.append(f"Planilha '{os.path.basename(file_path)}' carregada com sucesso.")
            self.log_text.append(f"Furos processados (até {max_furos} grupos) a partir das colunas 'furo_N_...'.")

        except Exception as e:
            QMessageBox.critical(self, "Erro de Leitura", f"Falha ao ler o arquivo: {e}\n\nVerifique o console para mais detalhes.")
            print(f"Erro detalhado ao ler Excel: {e}")
            import traceback
            traceback.print_exc()
    
    def clear_excel_data(self):
        self.excel_df = pd.DataFrame(columns=self.colunas_df); self.file_label.setText("Nenhuma planilha selecionada"); self.update_table_display()

    def import_dxfs(self):
        if not self.project_directory:
            QMessageBox.warning(self, "Nenhum Projeto Ativo", "Inicie um projeto antes de importar arquivos DXF.")
            return

        file_paths, _ = QFileDialog.getOpenFileNames(self, "Selecionar Arquivos DXF", "", "DXF Files (*.dxf)")
        if not file_paths:
            return

        start_time = time.time()
        imported_count = 0
        for file_path in file_paths:
            largura, altura = get_dxf_bounding_box(file_path)

            if largura is not None and altura is not None:
                nome_arquivo = os.path.splitext(os.path.basename(file_path))[0]
                
                new_piece = { 
                    'nome_arquivo': nome_arquivo,
                    'forma': 'rectangle', 
                    'forma': 'dxf_shape',
                    'espessura': 0.0, # Padrão
                    'qtd': 1, # Padrão
                    'largura': round(largura, 2),
                    'altura': round(altura, 2),
                    'diametro': 0.0, 'rt_base': 0.0, 'rt_height': 0.0,
                    'trapezoid_large_base': 0.0, 'trapezoid_small_base': 0.0, 'trapezoid_height': 0.0,
                    'furos': [],
                    'dxf_path': file_path
                }
                self.manual_df = pd.concat([self.manual_df, pd.DataFrame([new_piece])], ignore_index=True)
                imported_count += 1
            else:
                self.log_text.append(f"AVISO: Não foi possível obter as dimensões do arquivo '{os.path.basename(file_path)}'. Pode estar vazio ou corrompido.")
        
        elapsed_time = time.time() - start_time
        self.log_text.append(f"--- {imported_count} arquivo(s) DXF importado(s) com sucesso. (Tempo: {elapsed_time:.2f}s) ---")
    
    def replicate_holes(self):
        try:
            if self.forma_combo.currentText() != 'rectangle': QMessageBox.warning(self, "Função Indisponível", "Replicação disponível apenas para Retângulos."); return
            largura, altura = float(self.largura_input.text().replace(',', '.')), float(self.altura_input.text().replace(',', '.'))
            diam, offset = float(self.rep_diam_input.text().replace(',', '.')), float(self.rep_offset_input.text().replace(',', '.'))
            if (offset * 2) >= largura or (offset * 2) >= altura: QMessageBox.warning(self, "Offset Inválido", "Offset excede as dimensões da peça."); return
            furos = [{'diam': diam, 'x': offset, 'y': offset}, {'diam': diam, 'x': largura - offset, 'y': offset}, {'diam': diam, 'x': largura - offset, 'y': altura - offset}, {'diam': diam, 'x': offset, 'y': altura - offset}]
            self.furos_atuais.extend(furos); self.update_furos_table()
        except ValueError: QMessageBox.critical(self, "Erro de Valor", "Largura, Altura, Diâmetro e Offset devem ser números válidos.")
    
    def update_dimension_fields(self, shape):
        shape = shape.lower()
        is_rect, is_circ, is_tri, is_trap = shape == 'rectangle', shape == 'circle', shape == 'right_triangle', shape == 'trapezoid'
        for w in self.largura_row + self.altura_row: w.setVisible(is_rect)
        for w in self.diametro_row: w.setVisible(is_circ)
        for w in self.rt_base_row + self.rt_height_row: w.setVisible(is_tri)
        for w in self.trap_large_base_row + self.trap_small_base_row + self.trap_height_row: w.setVisible(is_trap)
        self.rep_group.setEnabled(is_rect)
        if hasattr(self, 'manual_group'):
            self.manual_group.adjustSize()
            self.manual_group.updateGeometry()
        if hasattr(self, 'manual_scroll_area'):
            self.manual_scroll_area.widget().adjustSize()
            self.manual_scroll_area.ensureVisible(0, 0)
    
    def add_furo_temp(self):
        try:
            diam, pos_x, pos_y = float(self.diametro_furo_input.text().replace(',', '.')), float(self.pos_x_input.text().replace(',', '.')), float(self.pos_y_input.text().replace(',', '.'))
            if diam <= 0: QMessageBox.warning(self, "Valor Inválido", "Diâmetro do furo deve ser maior que zero."); return
            self.furos_atuais.append({'diam': diam, 'x': pos_x, 'y': pos_y}); self.update_furos_table()
            for field in [self.diametro_furo_input, self.pos_x_input, self.pos_y_input]: field.clear()
        except ValueError: QMessageBox.critical(self, "Erro de Valor", "Campos de furo devem ser números válidos.")
    
    def update_furos_table(self):
        self.furos_table.setRowCount(0); self.furos_table.setRowCount(len(self.furos_atuais))
        for i, furo in enumerate(self.furos_atuais):
            self.furos_table.setItem(i, 0, QTableWidgetItem(str(furo['diam'])))
            self.furos_table.setItem(i, 1, QTableWidgetItem(str(furo['x'])))
            self.furos_table.setItem(i, 2, QTableWidgetItem(str(furo['y'])))
            delete_btn = QPushButton("Excluir")
            delete_btn.clicked.connect(lambda _, r=i: self.delete_furo_temp(r))
            self.furos_table.setCellWidget(i, 3, delete_btn)
        self.furos_table.resizeColumnsToContents()
    
    def delete_furo_temp(self, row_index):
        if 0 <= row_index < len(self.furos_atuais):
            del self.furos_atuais[row_index]
            self.update_furos_table()
def main():
    app = QApplication(sys.argv)
    app_icon_path = first_existing_desktop_asset(
        "dbx-ly.ico",
        "dbx-ly.png",
        "dbx-ly.svg",
    )
    if app_icon_path:
        app.setWindowIcon(QIcon(app_icon_path))
    app.setStyleSheet(STYLE)
    window = MainWindow()
    window.showMaximized() 
    #window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
